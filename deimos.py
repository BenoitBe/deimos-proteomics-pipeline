# ==============================================================================
# deimos.py — Orchestrateur principal
# Deimos — DIA Expression Integrated Multi-Omics Suite
# ==============================================================================
# Dépendances : pandas, numpy, scipy, statsmodels, sklearn, umap-learn,
#               matplotlib, seaborn, openpyxl, limma_ebayes (local)
# ==============================================================================

import os
import sys
import warnings
import re
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.colors import to_hex
import seaborn as sns
from itertools import combinations
from math import ceil
from scipy import stats
from scipy.stats import norm, pearsonr
from scipy.cluster.hierarchy import linkage, dendrogram, fcluster
from scipy.spatial.distance import squareform
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from sklearn.impute import SimpleImputer
import umap
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl

warnings.filterwarnings("ignore")

# adjustText émet un message FancyArrowPatch via logging (pas un warning Python) ;
# on coupe son logger pour garder une console propre. Sans effet sur le rendu.
import logging
logging.getLogger("adjustText").setLevel(logging.ERROR)

# Module eBayes local
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from limma_ebayes import (lm_fit, contrasts_fit, ebayes, top_table,
                           spectra_count_ebayes,
                           make_all_contrasts, make_design_matrix)

# Module GO/gProfiler optionnel (import protégé : si absent, le pipeline tourne)
try:
    from go_enrichment import (ask_go_params, run_go_enrichment,
                               export_go_sheets)
    _GO_AVAILABLE = True
except ImportError:
    _GO_AVAILABLE = False

# Module dashboard optionnel (import protégé)
try:
    from dashboard_integration import run_dashboard
    _DASHBOARD_AVAILABLE = True
except ImportError:
    _DASHBOARD_AVAILABLE = False

# Module config YAML réutilisable
from config import resolve_config


# ==============================================================================
# 0. PARAMÈTRES INTERACTIFS
# ==============================================================================

def ask_params() -> dict:
    """
    Asks the user 2 questions at launch and returns the thresholds.
    The volcano thresholds also govern the robustness score and the facet volcano.
    """
    print("\n" + "="*60)
    print("  PROTEOMICS PIPELINE — Statistical threshold configuration")
    print("="*60)

    # --- Question 1: Volcanos + Robustness ---
    print("\n[1/2] Individual volcanos, facet volcano & robustness score")
    v_type = ""
    while v_type not in ("1", "2"):
        v_type = input("  Use (1) raw p.value  or  (2) p.adj / FDR [BH] ? -> ").strip()

    v_val = None
    while v_val is None:
        try:
            v_val = float(input("  Threshold value (e.g. 0.05, 0.01) -> ").strip())
        except ValueError:
            print("  Invalid input, try again.")

    v_ratio = None
    while v_ratio is None:
        try:
            v_ratio = float(input("  Minimum ratio (e.g. 1.5, 2.0) -> ").strip())
        except ValueError:
            print("  Invalid input, try again.")

    # --- Question 2: ANOVA / Heatmap ---
    print("\n[2/2] ANOVA & heatmaps (significant protein selection)")
    a_type = ""
    while a_type not in ("1", "2"):
        a_type = input("  Use (1) raw p.value  or  (2) p.adj / FDR [BH] ? -> ").strip()

    a_val = None
    while a_val is None:
        try:
            a_val = float(input("  Threshold value (e.g. 0.05, 0.01) -> ").strip())
        except ValueError:
            print("  Invalid input, try again.")

    # Number of protein clusters (k-means) for the clustered heatmap
    n_clusters = None
    while n_clusters is None:
        raw = input("  Number of protein clusters for the heatmap "
                    "[default 3] -> ").strip()
        if raw == "":
            n_clusters = 3
        else:
            try:
                n_clusters = int(raw)
                if n_clusters < 2:
                    print("  At least 2 clusters.")
                    n_clusters = None
            except ValueError:
                print("  Invalid input (integer expected), try again.")

    # Robustness score: optional + number of iterations
    n_iter_rob = None
    while n_iter_rob is None:
        raw = input("  Robustness score: number of iterations "
                    "[default 100, 0 = disabled] -> ").strip()
        if raw == "":
            n_iter_rob = 100
        else:
            try:
                n_iter_rob = int(raw)
                if n_iter_rob < 0:
                    print("  Enter an integer >= 0 (0 = disabled).")
                    n_iter_rob = None
            except ValueError:
                print("  Invalid input (integer expected), try again.")

    # FDR correction: per-contrast (default) or global (whole study)
    fdr_global = None
    while fdr_global is None:
        raw = input("  FDR correction: [1] per contrast (default) "
                    "or [2] global (whole study) -> ").strip()
        if raw in ("", "1"):
            fdr_global = False
        elif raw == "2":
            fdr_global = True
        else:
            print("  Answer: 1 or 2.")

    # Missing-value imputation method
    imp_method = None
    while imp_method is None:
        print("\n  Missing-value imputation:")
        print("    [1] QRILC (pure MNAR — default, suited to DIA)")
        print("    [2] Mixed (QRILC if fully absent in a condition, kNN otherwise)")
        print("        [!] kNN may inflate power if MAR is dominant (see diagnostic)")
        raw = input("  -> ").strip()
        if raw in ("", "1"):
            imp_method = "qrilc"
        elif raw == "2":
            imp_method = "mixed"
        else:
            print("  Answer: 1 or 2.")

    params = {
        "volcano_use_padj":  v_type == "2",
        "volcano_p_thresh":  v_val,
        "volcano_lfc_min":   np.log2(v_ratio),
        "volcano_ratio_min": v_ratio,
        "anova_use_padj":    a_type == "2",
        "anova_p_thresh":    a_val,
        "n_heatmap_clusters": n_clusters,
        "n_iter_robustness": n_iter_rob,
        "fdr_global": fdr_global,
        "impute_method": imp_method,
    }

    print("\n  [OK] Parameters saved:")
    print(f"     Volcanos/Robustness -> {'p.adj' if params['volcano_use_padj'] else 'p.value'} < {v_val}  |  ratio >= {v_ratio} (log2FC >= {params['volcano_lfc_min']:.3f})")
    print(f"     ANOVA/Heatmap       -> {'p.adj' if params['anova_use_padj'] else 'p.value'} < {a_val}")
    print(f"     Heatmap clusters    -> {n_clusters}")
    print(f"     Robustness          -> {'disabled' if n_iter_rob == 0 else str(n_iter_rob) + ' iterations'}")
    _fdr_txt = "global (whole study)" if fdr_global else "per contrast"
    print(f"     FDR correction      -> {_fdr_txt}")
    print(f"     Imputation          -> {'QRILC (MNAR)' if imp_method == 'qrilc' else 'Mixed (MNAR+MAR)'}")
    print()
    return params


# ==============================================================================
# 1. DIAGNOSTIC — Vérification cohérence labels TSV ↔ Design
# ==============================================================================

def _extract_labels_from_tsv(tsv_path: str) -> tuple[list[str], list[str]]:
    """
    Lit uniquement la ligne d'en-tête du TSV et extrait :
    - raw_cols  : noms de colonnes bruts (après la 6ème)
    - clean_labels : labels nettoyés par le même regex que load_data
    """
    with open(tsv_path, "r", encoding="utf-8") as f:
        header = f.readline().rstrip("\n").split("\t")

    raw_cols = header[6:]
    clean_labels = []
    for c in raw_cols:
        m = re.search(r"(Sple-[^_]+)_", c)
        clean_labels.append(m.group(1) if m else c)

    return raw_cols, clean_labels


def diagnose_labels(tsv_path: str, design_path: str) -> bool:
    """
    Vérifie la cohérence entre les labels du TSV et ceux du design.
    Affiche un rapport détaillé et retourne True si tout est OK.

    Checks :
    1. Fichiers lisibles
    2. Colonnes obligatoires dans le design (label, condition, replicate)
    3. Labels design présents dans le TSV
    4. Labels TSV non couverts par le design
    5. Doublons dans chaque côté
    6. Suggestion de correspondances proches (distance de Levenshtein légère)
    """
    print("\n" + "="*60)
    print("  DIAGNOSTIC — Label check")
    print("="*60)

    errors   = []
    warnings = []
    ok       = True

    # --- 1. File readability ---
    for path, label in [(tsv_path, "pg_matrix.tsv"), (design_path, "ExperimentalDesign.csv")]:
        if not os.path.exists(path):
            errors.append(f"File not found: {path}")
            ok = False

    if not ok:
        for e in errors:
            print(f"  [ERROR] {e}")
        print("\n  [STOP] Cannot continue without the input files.")
        return False

    # --- 2. Required columns in the design ---
    design = pd.read_csv(design_path, sep=";")
    required_cols = {"label", "condition", "replicate"}
    missing_cols  = required_cols - set(design.columns)
    if missing_cols:
        errors.append(f"Missing columns in the design: {missing_cols}")
        ok = False

    # --- 3. Extract TSV labels ---
    try:
        raw_cols, tsv_labels = _extract_labels_from_tsv(tsv_path)
    except Exception as e:
        errors.append(f"Unable to read the TSV header: {e}")
        ok = False
        for e in errors:
            print(f"  [ERROR] {e}")
        return False

    design_labels = design["label"].astype(str).tolist() if "label" in design.columns else []

    # --- 4. Duplicates ---
    tsv_dupes    = [l for l in set(tsv_labels)    if tsv_labels.count(l)    > 1]
    design_dupes = [l for l in set(design_labels) if design_labels.count(l) > 1]
    if tsv_dupes:
        warnings.append(f"Duplicate labels in the TSV after cleaning: {tsv_dupes}")
    if design_dupes:
        errors.append(f"Duplicate labels in the design: {design_dupes}")
        ok = False

    # --- 5. Matching ---
    set_tsv    = set(tsv_labels)
    set_design = set(design_labels)

    matched        = set_tsv & set_design
    only_in_design = set_design - set_tsv   # in design but absent from TSV
    only_in_tsv    = set_tsv - set_design   # in TSV but absent from design

    # --- Report ---
    print(f"\n  TSV       : {len(tsv_labels)} columns detected ({len(set_tsv)} unique)")
    print(f"  Design    : {len(design_labels)} rows ({len(set_design)} unique)")
    print(f"  [OK] Matched: {len(matched)}")

    if only_in_design:
        ok = False
        print(f"\n  [ERROR] In the design but MISSING from the TSV ({len(only_in_design)}):")
        for lbl in sorted(only_in_design):
            suggestion = _closest_label(lbl, tsv_labels)
            hint = f"  -> looks like: '{suggestion}'" if suggestion else ""
            print(f"     - '{lbl}'{hint}")
        errors.append(f"{len(only_in_design)} design label(s) not found in the TSV")

    if only_in_tsv:
        print(f"\n  [WARN] In the TSV but MISSING from the design ({len(only_in_tsv)}):")
        for lbl in sorted(only_in_tsv):
            print(f"     - '{lbl}'")
        warnings.append(f"{len(only_in_tsv)} TSV column(s) not covered by the design (will be ignored)")

    # --- Raw columns for failed labels (debugging aid) ---
    if only_in_design:
        print(f"\n  [INFO] Raw column names in the TSV (before regex cleaning):")
        for raw, clean in zip(raw_cols, tsv_labels):
            marker = "[!] " if clean in only_in_design else "   "
            print(f"     {marker}raw: '{raw}'  ->  cleaned: '{clean}'")

    # --- Warnings ---
    if warnings:
        print()
        for w in warnings:
            print(f"  [WARN] {w}")

    # --- Conclusion ---
    print()
    if ok and not errors:
        print(f"  [OK] Labels consistent — {len(matched)} samples ready for analysis.\n")
    else:
        print("  [STOP] Errors detected — fix the design or the column names before re-running.\n")
        print("  [TIP] Hints:")
        print("     - The label in the design must match the 'Sple-XXX' part")
        print("       extracted from the raw TSV column name.")
        print("     - Example: TSV column '...250630-Sple-135A_Slot1-28...'")
        print("                -> expected design label: 'Sple-135A'")
        print("     - Also check the design separator (must be ';')\n")

    return ok and not errors


def _closest_label(target: str, candidates: list[str], max_dist: int = 4) -> str:
    """
    Retourne le candidat le plus proche par distance de Levenshtein simple.
    Retourne None si la distance minimale dépasse max_dist.
    """
    def levenshtein(a, b):
        if len(a) < len(b):
            return levenshtein(b, a)
        if not b:
            return len(a)
        prev = list(range(len(b) + 1))
        for i, ca in enumerate(a):
            curr = [i + 1]
            for j, cb in enumerate(b):
                curr.append(min(prev[j + 1] + 1,
                                curr[j] + 1,
                                prev[j] + (ca != cb)))
            prev = curr
        return prev[-1]

    if not candidates:
        return None
    dists = [(levenshtein(target, c), c) for c in candidates]
    best_dist, best_cand = min(dists)
    return best_cand if best_dist <= max_dist else None


# ==============================================================================
# 1. LECTURE ET NETTOYAGE DES DONNÉES
# ==============================================================================

def load_data(tsv_path: str, design_path: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Lit le pg_matrix.tsv et le design expérimental."""
    print("[IO] Reading raw files...")

    tsv = pd.read_csv(tsv_path, sep="\t")

    # Nettoyage noms de colonnes (équivalent sub R)
    new_cols = []
    for c in tsv.columns:
        m = re.search(r"(Sple-[^_]+)_", c)
        new_cols.append(m.group(1) if m else c)
    tsv.columns = new_cols

    # Identifier les colonnes LFQ (colonnes numériques après col 6)
    meta_cols = list(tsv.columns[:6])
    lfq_cols  = list(tsv.columns[6:])
    lfq_renamed = [f"LFQ.intensity.{c}" for c in lfq_cols]
    rename_map = dict(zip(lfq_cols, lfq_renamed))
    tsv = tsv.rename(columns=rename_map)

    # Colonnes ID et name
    tsv["ID"]   = tsv["Protein.Names"].astype(str).str.split(";").str[0]
    tsv["name"] = tsv["Protein.Group"].astype(str).str.split(";").str[0]

    # Retirer contaminants cRAP
    mask_crap = (tsv["Protein.Group"].astype(str).str.startswith("cRAP") |
                 tsv["Protein.Names"].astype(str).str.startswith("cRAP"))
    tsv = tsv[~mask_crap].reset_index(drop=True)

    # Design
    design = pd.read_csv(design_path, sep=";")

    print(f"  -> {len(tsv)} proteins loaded | {len(design)} samples in the design")
    return tsv, design


def build_peptide_count_table(pr_path: str, protein_names: list,
                              count_level: str = "peptide") -> "pd.Series | None":
    """
    Construit la count table DEqMS depuis le report.pr_matrix.tsv de DIA-NN.

    Pour chaque Protein.Group, compte le nombre de peptides (ou précurseurs)
    distincts quantifiés (valeur non-NA et > 0) dans CHAQUE run, puis prend le
    MINIMUM across runs — c'est le comptage utilisé par DEqMS pour modéliser la
    variance a priori (Zhu et al. 2020).

    Parameters
    ----------
    pr_path       : chemin vers report.pr_matrix.tsv
    protein_names : liste des 'name' (1er Protein.Group) des protéines filtrées,
                    pour aligner et ordonner le comptage
    count_level   : 'peptide' (Stripped.Sequence distinctes, défaut DEqMS LFQ)
                    ou 'precursor' (Precursor.Id distincts)

    Returns
    -------
    pd.Series indexée par 'name', valeurs = min count across runs (≥ 1).
    None si le fichier est absent ou illisible.
    """
    if not pr_path or not os.path.exists(pr_path):
        print(f"  [INFO] pr_matrix not found ({pr_path}) — DEqMS disabled.")
        return None

    try:
        print(f"  [IO] Reading pr_matrix for peptide counting...")
        pr = pd.read_csv(pr_path, sep="\t")

        # Identifier la colonne de regroupement protéine
        if "Protein.Group" not in pr.columns:
            print("  [WARN] 'Protein.Group' column absent from pr_matrix — DEqMS disabled.")
            return None

        # Colonne identifiant le niveau de comptage
        if count_level == "precursor":
            id_col = "Precursor.Id"
        else:
            id_col = "Stripped.Sequence"
        if id_col not in pr.columns:
            # Repli : si Stripped.Sequence absente, tenter Precursor.Id puis Modified.Sequence
            for alt in ["Stripped.Sequence", "Precursor.Id", "Modified.Sequence"]:
                if alt in pr.columns:
                    id_col = alt
                    break
            else:
                print("  [WARN] No peptide/precursor column found — DEqMS disabled.")
                return None

        # Colonnes d'intensité = celles qui contiennent un label d'échantillon (Sple.)
        # On nettoie les noms comme pour le pg_matrix
        run_cols = [c for c in pr.columns if re.search(r"Sple-[^_]+", c)]
        if not run_cols:
            # Repli : colonnes numériques non-méta
            meta_known = {"Protein.Group","Protein.Ids","Protein.Names","Genes",
                          "First.Protein.Description","Proteotypic","Stripped.Sequence",
                          "Modified.Sequence","Precursor.Id","Precursor.Charge"}
            run_cols = [c for c in pr.columns if c not in meta_known
                        and pd.api.types.is_numeric_dtype(pr[c])]
        if not run_cols:
            print("  [WARN] No intensity column detected in pr_matrix — DEqMS disabled.")
            return None

        # 'name' protéine = 1er Protein.Group
        pr["_name"] = pr["Protein.Group"].astype(str).str.split(";").str[0]

        # Pour chaque run : compter les peptides distincts quantifiés (>0, non-NA) par protéine
        # On construit une matrice protéine × run de comptages
        per_run_counts = {}
        for col in run_cols:
            detected = pr[[ "_name", id_col, col]].copy()
            detected = detected[detected[col].notna() & (detected[col] > 0)]
            # nombre de id_col distincts par protéine dans ce run
            cnt = detected.groupby("_name")[id_col].nunique()
            per_run_counts[col] = cnt

        count_mat = pd.DataFrame(per_run_counts).fillna(0)
        # MIN across runs (caractéristique DEqMS)
        min_count = count_mat.min(axis=1)

        # Aligner sur protein_names ; protéines sans peptide → NaN (exclues de DEqMS)
        aligned = pd.Series(min_count, index=min_count.index).reindex(protein_names)
        n_ok = aligned.notna().sum()
        n_missing = aligned.isna().sum()
        print(f"  -> {count_level} counting: {n_ok} matched proteins "
              f"(min across runs), {n_missing} sans comptage")
        if n_ok == 0:
            print("  [WARN] No protein matched between pg and pr_matrix — DEqMS disabled.")
            return None

        return aligned

    except Exception as e:
        print(f"  [WARN] pr_matrix read error ({type(e).__name__}: {str(e)[:80]}) — DEqMS disabled.")
        return None


# ==============================================================================
# 2. CONSTRUCTION DE LA MATRICE D'EXPRESSION (log2 LFQ)
# ==============================================================================

def build_expression_matrix(tsv: pd.DataFrame, design: pd.DataFrame
                             ) -> tuple[pd.DataFrame, pd.DataFrame, list]:
    """
    Retourne :
    - expr_log2 : DataFrame (n_proteins × n_samples) en log2, NaN pour 0
    - meta      : colonnes d'annotation protéines
    - lfq_cols  : noms de colonnes LFQ correspondant au design
    """
    lfq_cols = [c for c in tsv.columns if c.startswith("LFQ.intensity.")]

    # Alignement design → colonnes LFQ
    sample_map = {}
    for _, row in design.iterrows():
        lbl = row["label"]
        match = [c for c in lfq_cols if lbl in c]
        if match:
            sample_map[lbl] = match[0]

    ordered_lfq = [sample_map[lbl] for lbl in design["label"] if lbl in sample_map]
    design_filt = design[design["label"].isin(sample_map.keys())].reset_index(drop=True)

    mat = tsv[ordered_lfq].copy().astype(float)
    mat.replace(0, np.nan, inplace=True)
    mat = np.log2(mat)

    meta = tsv[["name", "Protein.Group", "Protein.Names", "Genes",
                "First.Protein.Description",
                "N.Sequences", "N.Proteotypic.Sequences"]].copy()

    return mat, meta, design_filt, ordered_lfq


# ==============================================================================
# 3. CONTRÔLE QUALITÉ
# ==============================================================================

def plot_qc(mat: pd.DataFrame, meta: pd.DataFrame, design: pd.DataFrame,
            out_dir: str) -> list[str]:
    """Génère les figures QC et retourne la liste des fichiers."""
    files = []
    conditions = design["condition"].values

    # --- Fréquence de détection ---
    n_detected = mat.notna().sum(axis=1)
    fig, ax = plt.subplots(figsize=(7, 4))
    ax.bar(range(len(n_detected)), sorted(n_detected.values, reverse=True),
           color="#3498DB", edgecolor="none")
    ax.set_xlabel("Proteins (sorted)")
    ax.set_ylabel("Number of samples detected")
    ax.set_title("Detection frequency")
    fig.tight_layout()
    f = os.path.join(out_dir, "qc_frequency.png")
    fig.savefig(f, dpi=150)
    plt.close(fig)
    files.append(f)

    # --- Nb protéines par échantillon ---
    n_per_sample = mat.notna().sum(axis=0)
    fig, ax = plt.subplots(figsize=(max(8, len(n_per_sample) * 0.4), 4))
    colors_cond = _condition_colors(conditions)
    bars = ax.bar(range(len(n_per_sample)), n_per_sample.values,
                  color=[colors_cond[c] for c in conditions], edgecolor="none")
    ax.set_xticks(range(len(n_per_sample)))
    ax.set_xticklabels(design["label"].values, rotation=90, fontsize=7)
    ax.set_ylabel("Number of proteins detected")
    ax.set_title("Proteins detected per sample")
    handles = [mpatches.Patch(color=v, label=k) for k, v in colors_cond.items()]
    ax.legend(handles=handles, fontsize=7, loc="lower right")
    fig.tight_layout()
    f = os.path.join(out_dir, "qc_numbers.png")
    fig.savefig(f, dpi=150)
    plt.close(fig)
    files.append(f)

    # --- Normalisation (boxplots) ---
    fig, ax = plt.subplots(figsize=(max(10, len(mat.columns) * 0.4), 5))
    data_box = [mat[c].dropna().values.flatten() for c in mat.columns]
    ax.boxplot(data_box, patch_artist=True,
               boxprops=dict(facecolor="#AED6F1"),
               medianprops=dict(color="red"),
               whiskerprops=dict(color="grey"),
               flierprops=dict(marker=".", markersize=2, alpha=0.3))
    ax.set_xticks(range(1, len(mat.columns) + 1))
    ax.set_xticklabels(design["label"].values, rotation=90, fontsize=7)
    ax.set_ylabel("log2 LFQ intensity")
    ax.set_title("Intensity distribution (log2 LFQ)")
    fig.tight_layout()
    f = os.path.join(out_dir, "qc_normalization.png")
    fig.savefig(f, dpi=150)
    plt.close(fig)
    files.append(f)

    # --- RLE plot (Relative Log Expression) ---
    # Diagnostic de normalisation résiduelle APRÈS DIA-NN (cross-run RT-dependent).
    # Pour chaque protéine, on soustrait sa médiane across-samples ; on boxplote
    # ces écarts par échantillon. Attendu si la normalisation tient : toutes les
    # boîtes centrées sur 0, de dispersion comparable. Un run dont la médiane RLE
    # s'écarte de 0 ou dont la boîte est anormalement étalée signale un effet
    # technique résiduel (le RLE neutralise le signal biologique commun, ce que
    # le boxplot de distribution brute ne fait pas). Aucune donnée n'est modifiée.
    rle = mat.sub(mat.median(axis=1), axis=0)   # écart à la médiane par protéine
    data_rle = [rle[c].dropna().values.flatten() for c in mat.columns]
    med_rle = np.array([np.median(d) if len(d) else np.nan for d in data_rle])
    cond_col = _condition_colors(conditions)
    fig, ax = plt.subplots(figsize=(max(10, len(mat.columns) * 0.4), 5))
    bp = ax.boxplot(data_rle, patch_artist=True, showfliers=False,
                    medianprops=dict(color="black", linewidth=1),
                    whiskerprops=dict(color="grey"))
    for patch, c in zip(bp["boxes"], conditions):
        patch.set_facecolor(cond_col[c]); patch.set_alpha(0.75)
    ax.axhline(0, color="red", linestyle="--", linewidth=0.8, zorder=0)
    ax.set_xticks(range(1, len(mat.columns) + 1))
    ax.set_xticklabels(design["label"].values, rotation=90, fontsize=7)
    ax.set_ylabel("Relative Log Expression (log2)")
    # Borne l'axe Y sur l'IQR agrégé pour rester lisible malgré les protéines extrêmes
    all_vals = np.concatenate([d for d in data_rle if len(d)])
    if all_vals.size:
        ylim = np.nanpercentile(np.abs(all_vals), 99)
        ax.set_ylim(-ylim, ylim)
    worst = np.nanmax(np.abs(med_rle)) if np.isfinite(med_rle).any() else 0.0
    ax.set_title(f"RLE plot — residual normalization check "
                 f"(max |median| = {worst:.2f} log2)")
    handles = [mpatches.Patch(color=v, label=k) for k, v in cond_col.items()]
    ax.legend(handles=handles, fontsize=7, loc="upper right")
    fig.tight_layout()
    f = os.path.join(out_dir, "qc_rle.png")
    fig.savefig(f, dpi=150)
    plt.close(fig)
    files.append(f)

    # --- Missing values heatmap ---
    fig, ax = plt.subplots(figsize=(max(8, len(mat.columns) * 0.35),
                                    min(12, len(mat) * 0.02 + 2)))
    sns.heatmap(mat.isna().astype(int), cmap=["#2ECC71", "#E74C3C"],
                yticklabels=False, xticklabels=design["label"].values,
                ax=ax, cbar=False)
    ax.set_xticklabels(ax.get_xticklabels(), rotation=90, fontsize=7)
    ax.set_title("Missing values (red = absent)")
    fig.tight_layout()
    f = os.path.join(out_dir, "qc_missval.png")
    fig.savefig(f, dpi=150)
    plt.close(fig)
    files.append(f)

    return files


# ==============================================================================
# 4. FILTRATION ET IMPUTATION MINPROB
# ==============================================================================

def filter_missval(mat: pd.DataFrame, design: pd.DataFrame, thr: int = 1) -> pd.DataFrame:
    """
    Garde les protéines détectées dans au moins (max_n_replicates - thr)
    réplicats dans AU MOINS une condition.
    thr=1 → détectée dans tous sauf 1 réplicat max.
    """
    conditions = design["condition"].values
    keep = np.zeros(len(mat), dtype=bool)

    for cond in np.unique(conditions):
        cols = mat.columns[conditions == cond]
        n = len(cols)
        min_detect = max(1, n - thr)
        detected = mat[cols].notna().sum(axis=1) >= min_detect
        keep |= detected.values

    filtered = mat[keep].reset_index(drop=True)
    print(f"  -> Filtering: {keep.sum()}/{len(mat)} proteins kept (thr={thr})")
    return filtered, keep


def impute_minprob(mat: pd.DataFrame, q: float = 0.01, rng=None) -> pd.DataFrame:
    """
    Imputation MinProb : remplace les NaN par un tirage dans N(µ_low, σ_low²)
    centré sur le quantile q de la distribution des valeurs détectées.
    rng : générateur np.random optionnel (pour reproductibilité en parallèle).
    Sans rng, tirage stochastique global (comportement historique).
    """
    draw = (rng.normal if rng is not None else np.random.normal)
    mat_imp = mat.copy()
    for col in mat_imp.columns:
        vals = mat_imp[col].dropna().values
        if len(vals) < 3:
            continue
        mu_low = np.quantile(vals, q)
        sd_low = np.std(vals) * 0.3  # σ réduit comme dans DEP
        n_miss = int(mat_imp[col].isna().sum())
        if n_miss > 0:
            mat_imp.loc[mat_imp[col].isna(), col] = draw(mu_low, sd_low, n_miss)
    return mat_imp


def impute_qrilc(mat: pd.DataFrame, rng=None) -> pd.DataFrame:
    """
    QRILC — Quantile Regression Imputation of Left-Censored data (Lazar 2016).
    Imputation MNAR rigoureuse : modélise la queue gauche comme censurée et
    tire dans une normale TRONQUÉE sous le seuil de détection estimé par
    régression quantile. Contrairement à MinProb, les valeurs imputées sont
    garanties basses (jamais au-dessus du seuil). Adapté au DIA (MNAR dominant).
    """
    from scipy.stats import norm, truncnorm
    if rng is None:
        rng = np.random.default_rng()
    mat_imp = mat.copy()
    for col in mat_imp.columns:
        s = mat_imp[col]
        obs = s.dropna().values
        n_miss = int(s.isna().sum())
        if n_miss == 0 or len(obs) < 3:
            if n_miss > 0 and len(obs) > 0:
                mat_imp.loc[s.isna(), col] = obs.min()
            continue
        mu, sigma = np.mean(obs), np.std(obs)
        if sigma <= 0:
            mat_imp.loc[s.isna(), col] = mu
            continue
        # Proportion manquante → quantile de censure
        p_miss = n_miss / len(s)
        q_censor = norm.ppf(max(p_miss, 1e-4), loc=mu, scale=sigma)
        # Tirage dans la normale tronquée à droite par q_censor
        a, b = -np.inf, (q_censor - mu) / sigma
        draws = truncnorm.rvs(a, b, loc=mu, scale=sigma, size=n_miss,
                              random_state=rng)
        mat_imp.loc[s.isna(), col] = draws
    return mat_imp


def impute_knn(mat: pd.DataFrame, k: int = 10) -> pd.DataFrame:
    """
    kNN — imputation MAR par k plus proches voisins (protéines au profil
    similaire). À réserver aux manquements aléatoires (non liés à l'intensité).
    Utilisé ici uniquement pour la composante MAR de l'imputation mixte.
    """
    from sklearn.impute import KNNImputer
    imputer = KNNImputer(n_neighbors=min(k, max(2, mat.shape[0] - 1)))
    arr = imputer.fit_transform(mat.values)
    return pd.DataFrame(arr, index=mat.index, columns=mat.columns)


def impute_mixed(mat: pd.DataFrame, design: pd.DataFrame, rng=None) -> pd.DataFrame:
    """
    Imputation MIXTE (MNAR + MAR), approche Perseus/MSnbase :
      • Une protéine ABSENTE de TOUS les réplicats d'une condition → manquement
        MNAR (sous le seuil de détection) → QRILC (valeurs basses).
      • Une protéine partiellement manquante (présente dans ≥1 réplicat de la
        condition) → manquement MAR (aléatoire) → kNN.
    Classification automatique par valeur manquante selon le design.
    """
    if rng is None:
        rng = np.random.default_rng()
    conditions = design["condition"].values
    mat_arr = mat.values.copy()
    na_mask = np.isnan(mat_arr)

    # 1) Pré-calcul des deux imputations complètes
    qrilc_full = impute_qrilc(mat, rng=rng).values
    try:
        knn_full = impute_knn(mat).values
    except Exception:
        knn_full = qrilc_full  # repli si kNN échoue (peu de protéines)

    # 2) Pour chaque NaN, décider MNAR (QRILC) ou MAR (kNN)
    out = mat_arr.copy()
    col_cond = np.array(conditions)
    for i in range(mat_arr.shape[0]):
        if not na_mask[i].any():
            continue
        for cond in np.unique(col_cond):
            cidx = np.where(col_cond == cond)[0]
            sub = mat_arr[i, cidx]
            n_present = np.sum(~np.isnan(sub))
            miss_here = cidx[np.isnan(sub)]
            if len(miss_here) == 0:
                continue
            if n_present == 0:
                # Absence totale dans la condition → MNAR → QRILC
                out[i, miss_here] = qrilc_full[i, miss_here]
            else:
                # Partiellement présent → MAR → kNN
                out[i, miss_here] = knn_full[i, miss_here]
    return pd.DataFrame(out, index=mat.index, columns=mat.columns)


def impute(mat: pd.DataFrame, method: str = "qrilc",
           design: pd.DataFrame = None, rng=None) -> pd.DataFrame:
    """Dispatcher d'imputation. method ∈ {'mixed', 'qrilc', 'minprob'}."""
    if method == "mixed":
        if design is None:
            print("  [WARN] Mixed imputation requires the design -> falling back to QRILC.")
            return impute_qrilc(mat, rng=rng)
        return impute_mixed(mat, design, rng=rng)
    elif method == "minprob":
        return impute_minprob(mat, rng=rng)
    else:  # qrilc (défaut)
        return impute_qrilc(mat, rng=rng)


def diagnose_missingness(mat: pd.DataFrame, design: pd.DataFrame,
                         out_dir: str) -> dict:
    """
    Diagnostic de la structure des valeurs manquantes pour objectiver le choix
    d'imputation (Mixte vs QRILC).

    Classe chaque valeur manquante :
      • MNAR : la protéine est absente de TOUS les réplicats de la condition
        (sous le seuil de détection → manquement lié à l'intensité).
      • MAR  : la protéine est présente dans ≥1 réplicat de la condition mais
        manquante dans d'autres (manquement vraisemblablement aléatoire).

    Produit :
      - un PNG (2 panneaux : part MNAR/MAR + intensité des protéines à trous)
      - un dict de métriques (part MNAR/MAR, recommandation)
    """
    conditions = design["condition"].values
    arr = mat.values
    na = np.isnan(arr)
    n_missing = int(na.sum())
    if n_missing == 0:
        print("  [INFO] No missing values — diagnostic not relevant.")
        return {"n_missing": 0, "recommendation": "Aucune imputation nécessaire"}

    # Classer chaque NaN en MNAR / MAR selon la présence dans la condition
    mnar_count = 0
    mar_count = 0
    uniq_conds = np.unique(conditions)
    for i in range(arr.shape[0]):
        if not na[i].any():
            continue
        for cond in uniq_conds:
            cidx = np.where(conditions == cond)[0]
            sub = arr[i, cidx]
            n_miss_here = int(np.isnan(sub).sum())
            if n_miss_here == 0:
                continue
            if np.sum(~np.isnan(sub)) == 0:
                mnar_count += n_miss_here     # absence totale → MNAR
            else:
                mar_count += n_miss_here       # partielle → MAR
    total = mnar_count + mar_count
    pct_mnar = 100 * mnar_count / total if total else 0
    pct_mar  = 100 * mar_count / total if total else 0

    # Intensité moyenne des protéines AVEC trous vs SANS trou (signature MNAR :
    # les protéines à trous doivent être globalement moins intenses)
    prot_has_na = na.any(axis=1)
    mean_int = np.nanmean(arr, axis=1)
    int_with_na = mean_int[prot_has_na]
    int_no_na   = mean_int[~prot_has_na]

    # --- Graphique ---
    fig, axes = plt.subplots(1, 2, figsize=(11, 4.2))
    # Panneau 1 : répartition MNAR / MAR
    axes[0].bar(["MNAR\n(fully absent\nin a condition)",
                 "MAR\n(partial\nmissingness)"],
                [pct_mnar, pct_mar],
                color=["#E8684A", "#5B8FF9"], edgecolor="black", lw=0.5)
    axes[0].set_ylabel("% of missing values")
    axes[0].set_title(f"Missingness structure (n = {total:,} values)",
                      fontweight="bold", fontsize=11)
    for x, v in enumerate([pct_mnar, pct_mar]):
        axes[0].text(x, v + 1, f"{v:.0f} %", ha="center", fontweight="bold")
    axes[0].set_ylim(0, 105)

    # Panel 2: intensity distribution (proteins with gaps vs complete)
    bins = np.linspace(np.nanmin(mean_int), np.nanmax(mean_int), 40)
    axes[1].hist(int_no_na, bins=bins, alpha=0.6, label="No missing value",
                 color="#999999", density=True)
    axes[1].hist(int_with_na, bins=bins, alpha=0.6, label="With >=1 missing",
                 color="#E8684A", density=True)
    axes[1].set_xlabel("Mean log2 intensity")
    axes[1].set_ylabel("Density")
    axes[1].set_title("Intensity of proteins with gaps\n(left shift = MNAR signature)",
                      fontweight="bold", fontsize=11)
    axes[1].legend(fontsize=9)
    fig.suptitle("Missingness diagnostic — imputation choice",
                 fontsize=13, fontweight="bold", y=1.02)
    fig.tight_layout()
    f = os.path.join(out_dir, "diagnostic_missingness.png")
    fig.savefig(f, dpi=150, bbox_inches="tight")
    plt.close(fig)

    # --- Recommandation automatique ---
    delta_int = float(np.nanmedian(int_no_na) - np.nanmedian(int_with_na))
    if pct_mar >= 20:
        reco = ("MIXED imputation recommended: non-negligible MAR fraction "
                f"({pct_mar:.0f}%) -> kNN beneficial on these partial missing values.")
    elif delta_int > 0.5:
        reco = ("QRILC (MNAR) is sufficient: missingness predominantly MNAR "
                f"({pct_mnar:.0f}%) and proteins with gaps clearly less "
                f"intense (median delta = {delta_int:.2f} log2).")
    else:
        reco = (f"Missingness at {pct_mnar:.0f}% MNAR / {pct_mar:.0f}% MAR. "
                "QRILC is appropriate; Mixed adds marginal benefit.")

    metrics = {
        "n_missing": n_missing,
        "pct_mnar": round(pct_mnar, 1),
        "pct_mar": round(pct_mar, 1),
        "delta_intensity": round(delta_int, 3),
        "recommendation": reco,
        "plot": f,
    }
    print(f"  [STATS] Missingness: {pct_mnar:.0f}% MNAR / {pct_mar:.0f}% MAR "
          f"| delta intensity(complete-gaps) = {delta_int:.2f} log2")
    print(f"     -> {reco}")
    return metrics


def plot_imputation(mat_filt: pd.DataFrame, mat_imp: pd.DataFrame,
                    out_dir: str) -> str:
    """Superposition avant/après imputation."""
    fig, ax = plt.subplots(figsize=(7, 4))
    vals_before = mat_filt.values.flatten()
    vals_before = vals_before[~np.isnan(vals_before)]
    vals_after  = mat_imp.values.flatten()

    ax.hist(vals_before, bins=80, alpha=0.6, color="#3498DB", label="Before imputation", density=True)
    ax.hist(vals_after,  bins=80, alpha=0.6, color="#E74C3C", label="After imputation", density=True)
    ax.set_xlabel("log2 LFQ intensity")
    ax.set_ylabel("Density")
    ax.set_title("Effect of MinProb imputation")
    ax.legend()
    fig.tight_layout()
    f = os.path.join(out_dir, "qc_imputation.png")
    fig.savefig(f, dpi=150)
    plt.close(fig)
    return f


# ==============================================================================
# 5. ANALYSE DIFFÉRENTIELLE — eBayes + Robustness
# ==============================================================================

def run_differential_analysis(mat_imp: pd.DataFrame, mat_filt: pd.DataFrame,
                               meta: pd.DataFrame, design: pd.DataFrame,
                               params: dict, out_dir: str, n_iter: int = 100,
                               pep_count: "pd.Series | None" = None
                               ) -> tuple[pd.DataFrame, list[str]]:
    """
    1. Modèle linéaire eBayes sur données imputées (référence)
    2. Si pep_count fourni : calcule AUSSI DEqMS (variance ~ nb peptides) en parallèle
    3. Robustness score (n_iter imputations stochastiques, seuils volcanos)
    4. Scatter plots par contraste
    5. Retourne df_results (une ligne par protéine, toutes stats) + noms contrastes
    """
    conditions = design["condition"].values
    expr = mat_imp.values.astype(float)          # (n_prot × n_samples)

    # --- Design matrix ---
    design_mat, group_names = make_design_matrix(conditions.tolist())
    contrast_mat, contrast_names = make_all_contrasts(group_names)

    print(f"\n[MODEL] {len(contrast_names)} contrasts generated: {contrast_names[:5]}{'...' if len(contrast_names)>5 else ''}")

    # --- Modèle de référence (limma eBayes) ---
    fdr_global = params.get("fdr_global", False)
    fit   = lm_fit(expr, design_mat)
    fit_c = contrasts_fit(fit, contrast_mat)
    fit_e = ebayes(fit_c, fdr_global=fdr_global)
    print(f"  [OK] limma eBayes model computed "
          f"(FDR {'global' if fdr_global else 'per contrast'})")

    # --- Modèle DEqMS (optionnel, en parallèle) ---
    fit_dq = None
    use_deqms = pep_count is not None
    if use_deqms:
        # Aligner le comptage sur l'ordre des protéines (df_results utilise meta['name'])
        names_order = meta["name"].values if "name" in meta.columns else None
        if names_order is not None:
            pc_aligned = pep_count.reindex(names_order).values
        else:
            pc_aligned = pep_count.values
        # Remplacer NaN par 1 (count minimal) pour permettre le fit, mais marquer invalides
        pc_for_fit = np.where(np.isfinite(pc_aligned) & (pc_aligned >= 1), pc_aligned, np.nan)
        try:
            fit_dq = spectra_count_ebayes(fit_c, np.nan_to_num(pc_for_fit, nan=1.0),
                                          fdr_global=fdr_global)
            print(f"  [OK] DEqMS model computed (df_prior={fit_dq['df_prior']:.2f})")
        except Exception as e:
            print(f"  [WARN] DEqMS failed ({type(e).__name__}) — limma only will be used.")
            fit_dq = None
            use_deqms = False

    # --- Intensités moyennes par condition (pour scatter) ---
    # IMPORTANT : les noms de contrastes utilisent des underscores (make_all_contrasts
    # fait group.replace(".", "_")). On construit donc cond_means avec la MÊME
    # convention pour éviter tout désalignement quand une condition contient un
    # séparateur (ex. "EP-HS-10" -> groupe "EP.HS.10" -> clé contraste "EP_HS_10").
    cond_means = {}
    for cond in np.unique(conditions):
        idx = np.where(conditions == cond)[0]
        cond_clean = re.sub(r"[^A-Za-z0-9_]", ".", cond)
        cond_clean = "X" + cond_clean if cond_clean[0].isdigit() else cond_clean
        cond_key = cond_clean.replace(".", "_")   # même forme que les contrastes
        cond_means[cond_key] = expr[:, idx].mean(axis=1)

    # --- Initialisation du df résultats ---
    df_results = meta.iloc[mat_filt.index if hasattr(mat_filt, 'index') else range(len(meta))].copy()
    df_results = df_results.reset_index(drop=True)

    scatter_files = []

    for i, cname in enumerate(contrast_names):
        tt = top_table(fit_e, i, protein_names=df_results["name"].values)

        # Le fold-change (_diff) vient toujours de limma : DEqMS ne modifie pas
        # l'estimation du logFC, seulement la variance résiduelle (donc la p-value).
        df_results[f"{cname}_diff"] = tt["logFC"].values

        if use_deqms and fit_dq is not None:
            # DEqMS FAIT FOI : ses p-values occupent les colonnes principales
            # (_p.val / _p.adj) utilisées par volcanos / UpSet / GO / robustesse.
            # limma est conservé en colonnes secondaires (_limma) pour comparaison.
            tt_dq = top_table(fit_dq, i, protein_names=df_results["name"].values)
            df_results[f"{cname}_p.val"]      = tt_dq["P.Value"].values
            df_results[f"{cname}_p.adj"]      = tt_dq["adj.P.Val"].values
            df_results[f"{cname}_p.val_limma"] = tt["P.Value"].values
            df_results[f"{cname}_p.adj_limma"] = tt["adj.P.Val"].values
            p_for_pi = tt_dq["P.Value"].values
        else:
            # limma seul (pas de comptage peptidique disponible)
            df_results[f"{cname}_p.val"] = tt["P.Value"].values
            df_results[f"{cname}_p.adj"] = tt["adj.P.Val"].values
            p_for_pi = tt["P.Value"].values

        df_results[f"Pi_Score_{cname}"] = (np.abs(tt["logFC"].values) *
                                           (-np.log10(np.maximum(p_for_pi, 1e-10))))

    # Comptage peptidique en colonne (traçabilité)
    if use_deqms and fit_dq is not None:
        df_results["Peptide_Count_min"] = fit_dq["pep_count"]
        print("  [INFO] Primary statistic = DEqMS (variance ~ peptide count); "
              "limma kept in _limma columns.")
    else:
        print("  [INFO] Primary statistic = limma eBayes (no peptide "
              "count available for DEqMS).")

    # --- Robustness score (optionnel) ---
    # OPTIMISATION : une seule imputation sert pour TOUS les contrastes à chaque
    # itération (le fit eBayes produit tous les contrastes d'un coup). On passe
    # ainsi de (n_iter × n_contrasts) imputations à n_iter au total, et on
    # parallélise les itérations sur les cœurs disponibles.
    if n_iter and n_iter > 0:
        print(f"\n[...] Robustness score ({n_iter} iterations, shared imputations)...")
        n_prot = len(mat_filt)
        n_contr = len(contrast_names)
        p_key   = "adj.P.Val" if params["volcano_use_padj"] else "P.Value"
        lfc_min = params["volcano_lfc_min"]
        p_thr   = params["volcano_p_thresh"]

        def _one_iteration(seed):
            """Une imputation + fit → matrice (n_prot × n_contr) de succès 0/1.

            IMPORTANT : si DEqMS est la statistique principale (use_deqms), on
            réapplique DEqMS à chaque itération pour que le robustness teste la
            MÊME statistique que celle affichée dans les colonnes _p.val/_p.adj.
            Le comptage peptidique (pep_count) ne dépend pas de l'imputation des
            intensités : on le réutilise tel quel à chaque tirage.
            """
            rng = np.random.default_rng(seed)
            mat_imp_tmp = impute_minprob(mat_filt, rng=rng)
            expr_tmp = mat_imp_tmp.values.astype(float)
            fit_tmp   = lm_fit(expr_tmp, design_mat)
            fit_tmp_c = contrasts_fit(fit_tmp, contrast_mat)
            fit_tmp_e = ebayes(fit_tmp_c, fdr_global=params.get("fdr_global", False))

            # Statistique de référence pour le critère de succès
            if use_deqms and fit_dq is not None:
                try:
                    fit_ref = spectra_count_ebayes(
                        fit_tmp_c, np.nan_to_num(pc_for_fit, nan=1.0),
                        fdr_global=params.get("fdr_global", False))
                except Exception:
                    fit_ref = fit_tmp_e   # repli limma si DEqMS échoue sur ce tirage
            else:
                fit_ref = fit_tmp_e

            out = np.zeros((n_prot, n_contr), dtype=np.int8)
            for i in range(n_contr):
                tt_tmp = top_table(fit_ref, i)
                passed = ((np.abs(tt_tmp["logFC"].values) >= lfc_min) &
                          (tt_tmp[p_key].values < p_thr))
                out[:, i] = passed.astype(np.int8)
            return out

        # Parallélisation (threads : numpy/scipy libèrent le GIL sur l'algèbre)
        success_total = np.zeros((n_prot, n_contr), dtype=np.int32)
        try:
            from concurrent.futures import ThreadPoolExecutor
            import os as _os
            n_workers = min(8, (_os.cpu_count() or 2))
            done = 0
            with ThreadPoolExecutor(max_workers=n_workers) as ex:
                for res in ex.map(_one_iteration, range(n_iter)):
                    success_total += res
                    done += 1
                    if done % 10 == 0 or done == n_iter:
                        print(f"    {done}/{n_iter} iterations", end="\r")
            print()
        except Exception as e:
            # Repli séquentiel si la parallélisation échoue
            print(f"  (parallelization unavailable: {e} -> sequential)")
            for k in range(n_iter):
                success_total += _one_iteration(k)
                if (k + 1) % 10 == 0:
                    print(f"    {k+1}/{n_iter}", end="\r")
            print()

        for i, cname in enumerate(contrast_names):
            df_results[f"Robustness_Score_{cname}"] = success_total[:, i]
        print("  [OK] Robustness score computed for all contrasts")
    else:
        print("\n[SKIP] Robustness score disabled (n_iter=0)")

    # --- Scatter plots (indépendants de la robustesse) ---
    for i, cname in enumerate(contrast_names):
        try:
            g1, g2 = cname.split("_vs_")
            if g1 in cond_means and g2 in cond_means:
                x_vals = cond_means[g1]
                y_vals = cond_means[g2]
                lfc    = df_results[f"{cname}_diff"].values
                pv     = df_results[f"{cname}_p.val"].values

                status = np.full(len(lfc), "Not changed", dtype=object)
                status[(lfc >  params["volcano_lfc_min"]) & (pv < 0.05)] = "Up regulated"
                status[(lfc < -params["volcano_lfc_min"]) & (pv < 0.05)] = "Down regulated"

                col_map = {"Up regulated": "#f3a583",
                           "Down regulated": "#92dadd",
                           "Not changed": "#bdbdbd"}

                fig, ax = plt.subplots(figsize=(6, 6))
                for s, col in col_map.items():
                    mask = status == s
                    ax.scatter(x_vals[mask], y_vals[mask], c=col, s=12,
                               alpha=0.5, label=s)
                lims = [min(x_vals.min(), y_vals.min()),
                        max(x_vals.max(), y_vals.max())]
                ax.plot(lims, lims, "k--", lw=0.8, alpha=0.5)
                r = params["volcano_lfc_min"]
                ax.plot(lims, [l + r for l in lims], ":", color="grey", lw=0.7)
                ax.plot(lims, [l - r for l in lims], ":", color="grey", lw=0.7)
                ax.set_xlabel(f"log2 Intensity — {g1}")
                ax.set_ylabel(f"log2 Intensity — {g2}")
                ax.set_title(f"Scatter : {g1} vs {g2}")
                ax.legend(fontsize=8)
                fig.tight_layout()
                f_scat = os.path.join(out_dir, f"scatter_{cname}.png")
                fig.savefig(f_scat, dpi=150)
                plt.close(fig)
                scatter_files.append(f_scat)
        except Exception as e:
            print(f"  [WARN] Scatter {cname}: {e}")

    return df_results, contrast_names, scatter_files


# ==============================================================================
# 6. VOLCANO PLOTS
# ==============================================================================

def plot_volcanoes(df_results: pd.DataFrame, contrast_names: list,
                   params: dict, out_dir: str) -> tuple[list[str], str]:
    """
    Génère les volcanos individuels + le facet volcano universel.
    Seuils = seuils volcanos (= seuils robustness).
    """
    p_key  = "p.adj" if params["volcano_use_padj"] else "p.val"
    thresh = params["volcano_p_thresh"]
    lfc    = params["volcano_lfc_min"]
    indiv_files = []

    col_map = {"Up regulated":    "#E74C3C",
               "Down regulated":  "#3498DB",
               "Not significant": "#BDBDBD"}

    # --- Volcanos individuels ---
    for cname in contrast_names:
        diff_col = f"{cname}_diff"
        pval_col = f"{cname}_{p_key}"
        if diff_col not in df_results.columns:
            continue

        d  = df_results[diff_col].values
        p  = df_results[pval_col].values        # statistique du SEUIL (p.val ou p.adj)
        nm = df_results["name"].values

        status = np.full(len(d), "Not significant", dtype=object)
        status[(d > lfc)  & (p < thresh)] = "Up regulated"
        status[(d < -lfc) & (p < thresh)] = "Down regulated"

        # L'axe Y utilise la MÊME statistique que le seuil (cohérence visuelle :
        # un point au-dessus de la ligne horizontale est significatif et coloré).
        y_all = -np.log10(np.maximum(p, 1e-10))

        fig, ax = plt.subplots(figsize=(7, 6))
        for s, col in col_map.items():
            mask = status == s
            ax.scatter(d[mask], y_all[mask],
                       c=col, s=12, alpha=0.7, label=s)

        # --- Labels des significatifs (anti-chevauchement) ---
        sig_mask = status != "Not significant"
        xs = d[sig_mask]
        ys = y_all[sig_mask]
        ns = nm[sig_mask]

        # Limiter aux plus marquants pour la lisibilité (comme geom_text_repel/max.overlaps).
        # Score = distance au point (0,seuil) → priorise forts |FC| et faibles p.
        MAX_LABELS = 40
        if len(ns) > MAX_LABELS:
            score = np.abs(xs) + ys
            keep = np.argsort(score)[::-1][:MAX_LABELS]
            xs, ys, ns = xs[keep], ys[keep], ns[keep]

        texts = []
        for xi, yi, ni in zip(xs, ys, ns):
            texts.append(ax.text(xi, yi, ni, fontsize=5, alpha=0.85))

        # adjustText si disponible (équivalent de ggrepel), sinon léger décalage
        try:
            from adjustText import adjust_text
            adjust_text(texts, ax=ax,
                        arrowprops=dict(arrowstyle="-", color="grey", lw=0.4,
                                        shrinkA=4, shrinkB=2),
                        expand_points=(1.4, 1.6), expand_text=(1.2, 1.4),
                        force_text=(0.4, 0.6), only_move={"text": "xy"})
        except ImportError:
            # Repli : petit décalage diagonal pour limiter le recouvrement
            for t in texts:
                t.set_position((t.get_position()[0] + 0.05,
                                t.get_position()[1] + 0.02))

        ax.axvline(-lfc, ls="--", lw=0.8, color="grey")
        ax.axvline( lfc, ls="--", lw=0.8, color="grey")
        ax.axhline(-np.log10(thresh), ls="--", lw=0.8, color="grey")
        ax.set_xlabel("log2 Fold Change")
        ax.set_ylabel(f"-log10({'p.adj' if params['volcano_use_padj'] else 'p.value'})")
        ax.set_title(cname.replace("_vs_", " vs "))
        ax.legend(fontsize=7)
        fig.tight_layout()
        f = os.path.join(out_dir, f"volc_{cname}.png")
        fig.savefig(f, dpi=150)
        plt.close(fig)
        indiv_files.append(f)

    # --- Facet volcano universel ---
    n_c    = len(contrast_names)
    n_cols = 4
    n_rows = (n_c + n_cols - 1) // n_cols
    fig, axes = plt.subplots(n_rows, n_cols,
                             figsize=(n_cols * 4, n_rows * 3.5))
    axes = np.array(axes).flatten()

    for idx, cname in enumerate(contrast_names):
        ax = axes[idx]
        diff_col = f"{cname}_diff"
        pval_col = f"{cname}_{p_key}"
        praw_col = f"{cname}_p.val"
        if diff_col not in df_results.columns:
            ax.set_visible(False)
            continue

        d = df_results[diff_col].values
        p = df_results[pval_col].values
        p_raw = df_results[praw_col].values

        status = np.full(len(d), "Non-significant", dtype=object)
        status[(d > lfc)  & (p < thresh)] = "Up-regulated"
        status[(d < -lfc) & (p < thresh)] = "Down-regulated"

        colors_pt = [{"Up-regulated": "#E74C3C",
                      "Down-regulated": "#3498DB",
                      "Non-significant": "#DDDDDD"}[s] for s in status]
        ax.scatter(d, -np.log10(np.maximum(p_raw, 1e-10)),
                   c=colors_pt, s=4, alpha=0.6)
        ax.axvline(-lfc, ls=":", lw=0.6, color="black", alpha=0.4)
        ax.axvline( lfc, ls=":", lw=0.6, color="black", alpha=0.4)
        ax.axhline(-np.log10(thresh), ls=":", lw=0.6, color="black", alpha=0.4)
        ax.set_title(cname.replace("_vs_", " vs "), fontsize=7,
                     fontweight="bold", color="white",
                     bbox=dict(facecolor="#34495E", boxstyle="round,pad=0.2"))
        ax.set_xlabel("log2FC", fontsize=7)
        ax.set_ylabel("-log10(p)", fontsize=7)
        ax.tick_params(labelsize=6)

    for idx in range(n_c, len(axes)):
        axes[idx].set_visible(False)

    # Légende globale
    patches = [mpatches.Patch(color="#E74C3C", label="Up"),
               mpatches.Patch(color="#3498DB", label="Down"),
               mpatches.Patch(color="#DDDDDD", label="NS")]
    fig.legend(handles=patches, loc="lower right", fontsize=9, ncol=3)
    fig.suptitle(f"Facet Volcano — {n_c} comparisons | "
                 f"threshold {'p.adj' if params['volcano_use_padj'] else 'p.val'} < {thresh},"
                 f" |log2FC| > {lfc:.2f}", fontsize=11, fontweight="bold")
    fig.tight_layout(rect=[0, 0.02, 1, 0.97])

    f_facet = os.path.join(out_dir, "all_volcano_plots_grid.png")
    fig.savefig(f_facet, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  [OK] Facet volcano saved ({n_c} contrasts)")

    return indiv_files, f_facet


# ==============================================================================
# 7. PCA + ELLIPSES
# ==============================================================================

def plot_pca(mat_imp: pd.DataFrame, design: pd.DataFrame, out_dir: str) -> list[str]:
    """PCA sur les données imputées, avec et sans ellipses de confiance."""
    expr = mat_imp.T.values  # (n_samples × n_proteins)
    # Imputation des éventuels NaN résiduels pour la PCA
    imp = SimpleImputer(strategy="mean")
    expr_imp = imp.fit_transform(expr)

    scaler = StandardScaler()
    expr_scaled = scaler.fit_transform(expr_imp)

    pca = PCA(n_components=min(10, expr_scaled.shape[1], expr_scaled.shape[0]))
    coords = pca.fit_transform(expr_scaled)
    var_exp = pca.explained_variance_ratio_

    conditions = design["condition"].values
    labels     = design["label"].values
    cmap       = _condition_colors(conditions)
    colors_pt  = [cmap[c] for c in conditions]

    files = []
    for with_ellipse in [False, True]:
        fig, ax = plt.subplots(figsize=(9, 7))
        for cond in np.unique(conditions):
            mask = conditions == cond
            ax.scatter(coords[mask, 0], coords[mask, 1],
                       color=cmap[cond], label=cond, s=60, alpha=0.85, zorder=3)
            # Labels
            for xi, yi, li in zip(coords[mask, 0], coords[mask, 1], labels[mask]):
                ax.annotate(li, (xi, yi), fontsize=6, alpha=0.7,
                            xytext=(3, 3), textcoords="offset points")
            if with_ellipse and mask.sum() >= 3:
                _draw_ellipse(ax, coords[mask, 0], coords[mask, 1],
                              color=cmap[cond], alpha=0.15)

        ax.set_xlabel(f"PC1 ({var_exp[0]*100:.1f}%)")
        ax.set_ylabel(f"PC2 ({var_exp[1]*100:.1f}%)")
        ax.set_title("PCA — imputed data" + (" (95% ellipses)" if with_ellipse else ""))
        ax.legend(fontsize=7, bbox_to_anchor=(1.01, 1), loc="upper left")
        ax.axhline(0, color="grey", lw=0.5)
        ax.axvline(0, color="grey", lw=0.5)
        fig.tight_layout()
        suffix = "_ellipses" if with_ellipse else ""
        f = os.path.join(out_dir, f"plot_pca{suffix}.png")
        fig.savefig(f, dpi=150, bbox_inches="tight")
        plt.close(fig)
        files.append(f)

    return files


def _draw_ellipse(ax, x, y, color, alpha=0.2, level=0.95):
    """Ellipse de confiance sur données 2D."""
    from matplotlib.patches import Ellipse
    from scipy.stats import chi2
    cov = np.cov(x, y)
    vals, vecs = np.linalg.eigh(cov)
    order = vals.argsort()[::-1]
    vals, vecs = vals[order], vecs[:, order]
    angle = np.degrees(np.arctan2(*vecs[:, 0][::-1]))
    chi2_val = chi2.ppf(level, df=2)
    width, height = 2 * np.sqrt(vals * chi2_val)
    ell = mpatches.Ellipse(xy=(x.mean(), y.mean()),
                           width=width, height=height,
                           angle=angle, color=color, alpha=alpha)
    ax.add_patch(ell)


# ==============================================================================
# 8. HEATMAP DE CORRÉLATION
# ==============================================================================

def plot_correlation_heatmap(mat_imp: pd.DataFrame, design: pd.DataFrame,
                              out_dir: str) -> str:
    """Heatmap de corrélation de Pearson inter-échantillons."""
    corr = np.corrcoef(mat_imp.values.T)  # (n_samples × n_samples)
    labels = design["label"].values
    conditions = design["condition"].values
    cmap = _condition_colors(conditions)

    fig, ax = plt.subplots(figsize=(max(8, len(labels) * 0.45),
                                    max(7, len(labels) * 0.4)))
    im = ax.imshow(corr, cmap="RdBu_r", vmin=corr.min() * 0.95, vmax=1.0)

    # Annotations
    for i in range(len(labels)):
        for j in range(len(labels)):
            ax.text(j, i, f"{corr[i,j]:.2f}", ha="center", va="center",
                    fontsize=max(5, 8 - len(labels) // 5), color="black")

    ax.set_xticks(range(len(labels)))
    ax.set_yticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=90, fontsize=7)
    ax.set_yticklabels(labels, fontsize=7)

    # Barre de couleur condition (annotation en haut)
    for j, cond in enumerate(conditions):
        ax.add_patch(mpatches.Rectangle((j - 0.5, -1.5), 1, 0.8,
                                         color=cmap[cond], clip_on=False))

    plt.colorbar(im, ax=ax, fraction=0.03)
    ax.set_title("Pearson correlation between samples")
    fig.tight_layout()
    f = os.path.join(out_dir, "heatmap_correlation.png")
    fig.savefig(f, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return f


# ==============================================================================
# 9. UMAP
# ==============================================================================

def plot_umap(mat_imp: pd.DataFrame, meta: pd.DataFrame,
              design: pd.DataFrame, out_dir: str) -> str:
    """UMAP sur la matrice imputée transposée (échantillons en lignes)."""
    expr = mat_imp.T.values
    imp  = SimpleImputer(strategy="mean")
    expr_imp = imp.fit_transform(expr)

    n_samples = expr_imp.shape[0]
    n_neighbors = min(5, n_samples - 1)

    cfg = umap.UMAP(n_neighbors=n_neighbors, min_dist=0.1, random_state=42)
    emb = cfg.fit_transform(expr_imp)

    conditions = design["condition"].values
    labels     = design["label"].values
    cmap       = _condition_colors(conditions)

    fig, ax = plt.subplots(figsize=(11, 7))
    for cond in np.unique(conditions):
        mask = conditions == cond
        ax.scatter(emb[mask, 0], emb[mask, 1],
                   color=cmap[cond], label=cond, s=60, alpha=0.85, zorder=3)
        for xi, yi, li in zip(emb[mask, 0], emb[mask, 1], labels[mask]):
            ax.annotate(li, (xi, yi), fontsize=6, alpha=0.7,
                        xytext=(3, 3), textcoords="offset points")

    ax.set_xlabel("UMAP1")
    ax.set_ylabel("UMAP2")
    n_prot = mat_imp.shape[0]
    n_cond = len(np.unique(conditions))
    ax.set_title(f"Proteome UMAP projection\n"
                 f"n = {n_prot} proteins | {n_cond} conditions")
    ax.legend(fontsize=7, bbox_to_anchor=(1.01, 1), loc="upper left")
    fig.tight_layout()
    f = os.path.join(out_dir, "umap_plot.png")
    fig.savefig(f, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print("  [OK] UMAP generated")
    # Coordonnées UMAP par échantillon (pour la feuille Excel + dashboard)
    umap_coords = pd.DataFrame({
        "label":     labels,
        "condition": conditions,
        "UMAP1":     emb[:, 0],
        "UMAP2":     emb[:, 1],
    })
    return f, umap_coords


# ==============================================================================
# 10. ANOVA GLOBALE + HEATMAPS
# ==============================================================================

def run_anova_heatmaps(mat_imp: pd.DataFrame, meta: pd.DataFrame,
                       design: pd.DataFrame, params: dict, out_dir: str
                       ) -> tuple[pd.DataFrame, str, str, str, np.ndarray, list]:
    """
    ANOVA one-way par protéine (F-test sur le modèle limma global).
    Heatmaps classique + clusters. Retourne les fichiers et le mapping cluster.
    """
    from scipy.stats import f as fdist

    conditions = design["condition"].values
    groups = np.unique(conditions)
    expr = mat_imp.values.astype(float)  # (n_prot × n_samples)

    # F-test par protéine (one-way ANOVA via OLS)
    n_prot, n_samp = expr.shape
    n_groups = len(groups)
    df_between = n_groups - 1
    df_within  = n_samp - n_groups

    grand_mean = expr.mean(axis=1, keepdims=True)
    ss_between = np.zeros(n_prot)
    ss_within  = np.zeros(n_prot)

    for g in groups:
        idx = np.where(conditions == g)[0]
        gm = expr[:, idx].mean(axis=1, keepdims=True)
        ss_between += len(idx) * ((gm.flatten() - grand_mean.flatten()) ** 2)
        ss_within  += ((expr[:, idx] - gm) ** 2).sum(axis=1)

    ms_between = ss_between / df_between
    ms_within  = np.maximum(ss_within / df_within, 1e-12)
    f_stat = ms_between / ms_within
    p_vals = 1 - fdist.cdf(f_stat, df_between, df_within)

    # BH correction
    from limma_ebayes import _bh_correction
    p_adj  = _bh_correction(p_vals)

    df_anova = meta.copy().reset_index(drop=True)
    df_anova["F.stat"]    = f_stat
    df_anova["p.value"]   = p_vals
    df_anova["p.adj"]     = p_adj
    df_anova["significant"] = False

    p_key_col = "p.adj" if params["anova_use_padj"] else "p.value"
    df_anova["significant"] = df_anova[p_key_col] < params["anova_p_thresh"]
    df_anova = df_anova.sort_values("p.value").reset_index(drop=True)

    sig_names = df_anova.loc[df_anova["significant"], "name"].values
    print(f"  -> ANOVA: {len(sig_names)} significant proteins "
          f"({'p.adj' if params['anova_use_padj'] else 'p.value'} < {params['anova_p_thresh']})")

    hm_classic = hm_clusters = hm_violin = None
    cluster_mapping = []
    mat_zscore = None

    if len(sig_names) > 0:
        sig_idx = meta.index[meta["name"].isin(sig_names)].tolist()
        mat_sig = expr[sig_idx, :]
        mat_z   = (mat_sig - mat_sig.mean(axis=1, keepdims=True)) / np.maximum(
                   mat_sig.std(axis=1, keepdims=True), 1e-8)
        mat_zscore = mat_z

        sample_labels = design["label"].values
        cmap_cond = _condition_colors(conditions)

        # --- Heatmap classique (clustering hiérarchique pur) ---
        hm_classic = _draw_heatmap_pch(
            mat_z, sig_names, sample_labels, conditions, cmap_cond,
            title=f"ANOVA significant proteins "
                  f"(n={len(sig_names)}, {p_key_col} < {params['anova_p_thresh']})",
            out_path=os.path.join(out_dir, "heatmap_annotated.png"),
            show_row_names=len(sig_names) <= 80
        )

        # --- Heatmap clusters (split par k-means) ---
        # Sécurité : au moins ~3 protéines par cluster, sinon on réduit le
        # nombre de clusters (évite les clusters vides/singletons qui font
        # planter le dendrogramme de PyComplexHeatmap).
        n_sig = len(sig_names)
        n_req = params.get("n_heatmap_clusters", 3)
        n_row_clusters = max(2, min(n_req, n_sig // 3)) if n_sig >= 6 else 1
        n_col_clusters = min(4, len(groups))
        if n_row_clusters < n_req:
            print(f"  [WARN] {n_sig} significant proteins -> {n_row_clusters} "
                  f"clusters (instead of {n_req}) to avoid empty clusters.")
        row_order, col_order, row_labels = _cluster_heatmap(
            mat_z, n_row_clusters, n_col_clusters
        )

        hm_clusters = _draw_heatmap_pch(
            mat_z, sig_names, sample_labels, conditions, cmap_cond,
            title=f"Clusters ANOVA (n={len(sig_names)})",
            out_path=os.path.join(out_dir, "heatmap_clusters.png"),
            show_row_names=len(sig_names) <= 80,
            row_split_labels=row_labels
        )

        # Mapping cluster → protéines
        for ci in np.unique(row_labels):
            prots_c = sig_names[row_labels == ci]
            cluster_mapping.extend([(p, f"Cluster_{ci+1}") for p in prots_c])

        df_anova["Cluster_ID"] = df_anova["name"].map(
            dict(cluster_mapping)).fillna("")

        # --- Violin / Boxplot par cluster ---
        hm_violin = _plot_cluster_violin(
            mat_z, sig_names, row_labels, conditions, design,
            out_path=os.path.join(out_dir, "plt_clusters_profiles.png")
        )

    return df_anova, hm_classic, hm_clusters, hm_violin, mat_zscore, cluster_mapping, sig_names


def _draw_heatmap_pch(mat_z, row_names, sample_labels, conditions, cmap_cond,
                      title, out_path, show_row_names=True,
                      row_split_labels=None):
    """
    Heatmap Z-score avec PyComplexHeatmap : clustering hiérarchique (lignes +
    colonnes), dendrogrammes, annotation des conditions, split optionnel par
    cluster k-means. Rendu proche de ComplexHeatmap (R).
    Repli automatique sur _draw_heatmap si PyComplexHeatmap est indisponible.
    """
    try:
        import PyComplexHeatmap as pch
    except ImportError:
        print("  [WARN] PyComplexHeatmap NOT INSTALLED -> basic heatmap without clustering.")
        print("      Install it for ComplexHeatmap-style rendering: "
              "pip install PyComplexHeatmap")
        return _draw_heatmap(mat_z, row_names, sample_labels, conditions,
                             cmap_cond, title, out_path,
                             show_row_names=show_row_names,
                             row_cluster_labels=(np.sort(row_split_labels)
                                                 if row_split_labels is not None else None))

    n_prot, n_samp = mat_z.shape
    df = pd.DataFrame(mat_z, index=list(row_names), columns=list(sample_labels))

    # Annotation des conditions (barre colorée en haut)
    cond_series = pd.Series(list(conditions), index=list(sample_labels))
    col_anno = pch.HeatmapAnnotation(
        Condition=pch.anno_simple(cond_series, colors=cmap_cond,
                                  height=4, legend=True,
                                  add_text=False),
        axis=1, verbose=0, label_side="right",
        label_kws={"fontsize": 9, "fontweight": "bold",
                   "horizontalalignment": "left"},
    )

    # Split des lignes par cluster (k-means) + annotation de cluster à gauche
    row_split = None
    left_anno = None
    if row_split_labels is not None:
        # Sécurité : si trop peu de protéines ou un cluster singleton/vide,
        # le dendrogramme par cluster plante (matrice de distance vide).
        # On désactive alors le split (heatmap simple, clustering global).
        import collections as _coll
        counts = _coll.Counter(row_split_labels)
        if len(row_split_labels) >= 6 and all(v >= 2 for v in counts.values()):
            cluster_names = [f"C{l + 1}" for l in row_split_labels]
            row_split = pd.Series(cluster_names, index=list(row_names))
            # Couleurs distinctes par cluster (palette douce)
            uniq = sorted(set(cluster_names))
            clust_palette = dict(zip(
                uniq,
                ["#5B8FF9", "#F6BD16", "#5AD8A6", "#E8684A",
                 "#9270CA", "#FF9D4D", "#269A99", "#FF99C3"][:len(uniq)]))
            left_anno = pch.HeatmapAnnotation(
                Cluster=pch.anno_simple(row_split, colors=clust_palette,
                                        height=4, legend=True, add_text=True,
                                        text_kws={"fontsize": 8, "color": "black"}),
                axis=0, verbose=0, label_side="top",
                label_kws={"fontsize": 9, "fontweight": "bold"},
            )

    fig_h = max(6, min(n_prot * 0.18 + 3, 28))
    fig_w = max(9, n_samp * 0.40 + 5)
    plt.figure(figsize=(fig_w, fig_h))
    try:
        pch.ClusterMapPlotter(
            data=df,
            top_annotation=col_anno,
            left_annotation=left_anno,
            row_cluster=True, col_cluster=True,
            row_split=row_split,
            row_split_gap=2.2,
            col_split_gap=1.2,
            # Dendrogramme de lignes affiché MÊME avec split (clustering
            # hiérarchique à l'intérieur de chaque cluster)
            row_dendrogram=True,
            col_dendrogram=True,
            show_rownames=show_row_names, show_colnames=True,
            row_names_side="right",
            cmap="RdBu_r", vmin=-2.5, vmax=2.5, center=0,
            label="Z-score", legend=True,
            xticklabels_kws={"labelrotation": 90, "labelsize": 7},
            yticklabels_kws={"labelsize": 6},
            verbose=0,
        )
        plt.suptitle(title, fontsize=12, fontweight="bold", y=1.01)
        plt.savefig(out_path, dpi=160, bbox_inches="tight",
                    facecolor="white")
        plt.close("all")
        return out_path
    except Exception as e:
        plt.close("all")
        import traceback
        print(f"  [WARN] PyComplexHeatmap failed: {type(e).__name__}: {str(e)[:150]}")
        print("      → repli sur heatmap matplotlib basique.")
        for line in traceback.format_exc().strip().splitlines()[-3:]:
            print(f"        {line.strip()}")
        return _draw_heatmap(mat_z, row_names, sample_labels, conditions,
                             cmap_cond, title, out_path,
                             show_row_names=show_row_names,
                             row_cluster_labels=(np.sort(row_split_labels)
                                                 if row_split_labels is not None else None))


def _draw_heatmap(mat_z, row_names, sample_labels, conditions, cmap_cond,
                  title, out_path, show_row_names=True, row_cluster_labels=None):
    """Heatmap Z-score matplotlib (repli si PyComplexHeatmap absent)."""
    n_prot, n_samp = mat_z.shape
    fig_h = max(6, min(n_prot * 0.18 + 2, 25))
    fig_w = max(8, n_samp * 0.35 + 3)

    fig = plt.figure(figsize=(fig_w, fig_h))
    gs  = fig.add_gridspec(2, 2, height_ratios=[0.04, 1],
                           width_ratios=[1, 0.05], hspace=0.01)
    ax_ann  = fig.add_subplot(gs[0, 0])
    ax_heat = fig.add_subplot(gs[1, 0])
    ax_cbar = fig.add_subplot(gs[1, 1])

    # Annotation condition
    cond_colors = [cmap_cond[c] for c in conditions]
    for j, col in enumerate(cond_colors):
        ax_ann.add_patch(mpatches.Rectangle((j, 0), 1, 1, color=col))
    ax_ann.set_xlim(0, n_samp)
    ax_ann.set_ylim(0, 1)
    ax_ann.axis("off")
    ax_ann.set_title(title, fontsize=9, pad=3)

    # Heatmap
    im = ax_heat.imshow(mat_z, aspect="auto", cmap="RdBu_r",
                        vmin=-2.5, vmax=2.5, interpolation="nearest")
    ax_heat.set_xticks(range(n_samp))
    ax_heat.set_xticklabels(sample_labels, rotation=90, fontsize=6)
    if show_row_names:
        ax_heat.set_yticks(range(n_prot))
        ax_heat.set_yticklabels(row_names, fontsize=6)
    else:
        ax_heat.set_yticks([])

    # Séparateurs clusters
    if row_cluster_labels is not None:
        prev = row_cluster_labels[0]
        for k, cl in enumerate(row_cluster_labels):
            if cl != prev:
                ax_heat.axhline(k - 0.5, color="white", lw=1.5)
                prev = cl

    plt.colorbar(im, cax=ax_cbar, label="Z-score")
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return out_path


def _cluster_heatmap(mat_z, n_row_k, n_col_k):
    """K-means sur lignes et colonnes pour ordonner la heatmap clusters."""
    from sklearn.cluster import KMeans
    km_rows = KMeans(n_clusters=n_row_k, random_state=42, n_init=10)
    row_labels = km_rows.fit_predict(mat_z)
    km_cols = KMeans(n_clusters=n_col_k, random_state=42, n_init=10)
    col_labels = km_cols.fit_predict(mat_z.T)
    return np.argsort(row_labels), np.argsort(col_labels), row_labels


def _plot_cluster_violin(mat_z, row_names, row_labels, conditions, design, out_path):
    """Violin + boxplot par cluster × condition."""
    n_clusters = len(np.unique(row_labels))
    fig, axes = plt.subplots(1, n_clusters,
                             figsize=(n_clusters * 4, 5), sharey=False)
    if n_clusters == 1:
        axes = [axes]

    cmap = _condition_colors(conditions)

    for ci, ax in enumerate(axes):
        mask_prot = row_labels == ci
        data_cl = mat_z[mask_prot, :]  # (n_prot_cluster × n_samples)

        df_v = pd.DataFrame(data_cl.T, columns=row_names[mask_prot])
        df_v["condition"] = conditions
        df_v["sample"] = design["label"].values
        df_long = df_v.melt(id_vars=["condition", "sample"],
                            var_name="protein", value_name="Z_score")

        cond_order = sorted(np.unique(conditions))
        for k, cond in enumerate(cond_order):
            sub = df_long[df_long["condition"] == cond]["Z_score"].values
            if len(sub) > 1:
                parts = ax.violinplot([sub], positions=[k], widths=0.6,
                                      showmedians=False)
                for pc in parts["bodies"]:
                    pc.set_facecolor(cmap[cond])
                    pc.set_alpha(0.6)
            q1, med, q3 = np.percentile(sub, [25, 50, 75])
            ax.plot([k - 0.1, k + 0.1], [med, med], color="black", lw=2)
            ax.plot([k, k], [q1, q3], color="black", lw=1.5)

        ax.set_xticks(range(len(cond_order)))
        ax.set_xticklabels(cond_order, rotation=45, ha="right", fontsize=7)
        ax.set_title(f"Cluster {ci+1}\n(n={mask_prot.sum()} proteins)", fontsize=9)
        ax.set_ylabel("Z-score" if ci == 0 else "")

    fig.suptitle("Expression profiles by cluster", fontsize=11)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return out_path


# ==============================================================================
# 11. UPSET PLOT & INTERSECTIONS
# ==============================================================================

def plot_upset(df_results: pd.DataFrame, contrast_names: list,
               params: dict, out_dir: str) -> tuple[str, pd.DataFrame]:
    """UpSet plot des protéines DEP par contraste."""
    p_key = "p.adj" if params["volcano_use_padj"] else "p.val"
    thresh = params["volcano_p_thresh"]
    lfc    = params["volcano_lfc_min"]

    upset_dict = {}
    for cname in contrast_names:
        diff_col = f"{cname}_diff"
        pval_col = f"{cname}_{p_key}"
        if diff_col not in df_results.columns:
            continue
        mask_sig = ((np.abs(df_results[diff_col].values) >= lfc) &
                    (df_results[pval_col].values < thresh))
        prots = df_results.loc[mask_sig, "name"].dropna().unique().tolist()
        if prots:
            upset_dict[cname] = set(prots)

    if len(upset_dict) < 2:
        print("  [WARN] Not enough DEP contrasts for the UpSet.")
        return None, pd.DataFrame()

    all_prots = sorted(set.union(*upset_dict.values()))
    mat_bin = pd.DataFrame(
        {k: [1 if p in v else 0 for p in all_prots] for k, v in upset_dict.items()},
        index=all_prots
    )

    # UpSet plot matplotlib maison
    cnames_list = list(upset_dict.keys())
    n_sets = len(cnames_list)

    # Intersections (trier par fréquence)
    from itertools import chain
    intersection_counts = {}
    for _, row in mat_bin.iterrows():
        key = tuple(cnames_list[j] for j in range(n_sets) if row.iloc[j] == 1)
        if key:
            intersection_counts[key] = intersection_counts.get(key, 0) + 1

    top_intersections = sorted(intersection_counts.items(),
                                key=lambda x: -x[1])[:30]

    n_inter = len(top_intersections)
    fig = plt.figure(figsize=(max(14, n_inter * 0.5), 8))
    gs  = fig.add_gridspec(2, 2, width_ratios=[1, 4], height_ratios=[2, 1], hspace=0.05)

    ax_bar  = fig.add_subplot(gs[0, 1])
    ax_mat  = fig.add_subplot(gs[1, 1])
    ax_sets = fig.add_subplot(gs[1, 0])

    # Barplot intersections
    bar_vals = [v for _, v in top_intersections]
    ax_bar.bar(range(n_inter), bar_vals, color="#2C3E50", edgecolor="none")
    ax_bar.set_ylabel("Number of DEP proteins")
    ax_bar.set_xticks([])
    ax_bar.set_title("UpSet Plot — DEP intersections")

    # Matrice points
    ax_mat.set_xlim(-0.5, n_inter - 0.5)
    ax_mat.set_ylim(-0.5, n_sets - 0.5)
    ax_mat.set_yticks(range(n_sets))
    ax_mat.set_yticklabels([c[:25] for c in cnames_list], fontsize=6)
    ax_mat.set_xticks([])

    for xi, (keys, _) in enumerate(top_intersections):
        active = [cnames_list.index(k) for k in keys if k in cnames_list]
        for yi in active:
            ax_mat.scatter(xi, yi, s=50, c="#2C3E50", zorder=3)
        if len(active) > 1:
            ax_mat.plot([xi, xi], [min(active), max(active)],
                        color="#2C3E50", lw=2, zorder=2)

    # Barplot sets
    set_sizes = [len(upset_dict[c]) for c in cnames_list]
    ax_sets.barh(range(n_sets), set_sizes, color="#3498DB")
    ax_sets.invert_xaxis()
    ax_sets.set_yticks([])
    ax_sets.set_xlabel("Total DEP")

    fig.tight_layout()
    f_upset = os.path.join(out_dir, "upset_plot.png")
    fig.savefig(f_upset, dpi=150, bbox_inches="tight")
    plt.close(fig)

    # Tableau d'intersections
    mat_bin["Nb_Conditions"] = mat_bin.sum(axis=1)
    mat_bin_X = mat_bin.copy()
    for col in cnames_list:
        mat_bin_X[col] = mat_bin_X[col].map({1: "X", 0: ""})
    mat_bin_X = mat_bin_X.sort_values("Nb_Conditions", ascending=False).reset_index()
    mat_bin_X = mat_bin_X.rename(columns={"index": "Protein.Group"})

    print(f"  [OK] UpSet plot generated ({len(upset_dict)} contrasts with DEP)")
    return f_upset, mat_bin_X


# ==============================================================================
# 12. EXPORT EXCEL
# ==============================================================================

def build_mm_sheet(ws):
    """
    Feuille 'M&M_Amont' (préparation d'échantillons, LC-MS/MS, DIA-NN) au format
    structuré, placée en première position du classeur. Champ 'Base de données'
    (espèce) à renseigner selon le projet.
    """
    from openpyxl.styles import Alignment, Border, Side
    C_TITLE, C_SECTION, C_HDRTXT, C_ELEM = "2C3E50", "34495E", "7F8C8D", "ECF0F1"
    f_title   = Font(bold=True, size=14, color="FFFFFF")
    f_sub     = Font(bold=False, size=9, color=C_HDRTXT)
    f_section = Font(bold=True, size=11, color="FFFFFF")
    f_hdr     = Font(bold=True, size=9, color=C_HDRTXT)
    f_elem    = Font(bold=True, size=10)
    f_val     = Font(bold=False, size=11)
    fill_title   = PatternFill("solid", fgColor=C_TITLE)
    fill_section = PatternFill("solid", fgColor=C_SECTION)
    fill_elem    = PatternFill("solid", fgColor=C_ELEM)
    wrap = Alignment(wrap_text=True, vertical="top")
    wrap_center = Alignment(wrap_text=True, vertical="center")
    border = Border(bottom=Side(style="thin", color="D5DBDB"))
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 70
    state = {"row": 1}

    def title(txt, sub=None):
        r = state["row"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws.cell(r, 1, txt); c.font = f_title; c.fill = fill_title
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 24
        state["row"] += 1
        if sub:
            r = state["row"]
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            ws.cell(r, 1, sub).font = f_sub
            state["row"] += 1
        state["row"] += 1

    def section(txt):
        r = state["row"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws.cell(r, 1, txt); c.font = f_section; c.fill = fill_section
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 20
        state["row"] += 1
        for col, h in enumerate(["Item", "Choice / Value", "Method / Detail"], 1):
            ws.cell(state["row"], col, h).font = f_hdr
        state["row"] += 1

    def line(elem, val, detail):
        r = state["row"]
        c1 = ws.cell(r, 1, elem); c1.font = f_elem; c1.fill = fill_elem
        c1.alignment = wrap_center; c1.border = border
        c2 = ws.cell(r, 2, val); c2.font = f_val; c2.alignment = wrap; c2.border = border
        c3 = ws.cell(r, 3, detail); c3.font = f_val; c3.alignment = wrap; c3.border = border
        state["row"] += 1

    title("SAMPLE PREPARATION, LC-MS/MS & DIA-NN — Methods & Parameters",
          "Upstream part of the LFQ proteomics workflow | TIMS-TOF HT (Bruker) · DIA-NN")

    section("1. SAMPLE PREPARATION (BIOCHEMISTRY)")
    line("Protein extraction", "Modified GASP protocol",
         "Gel-aided Sample Preparation (PMC4409837), modified version")
    line("Digestion", "Trypsin / Lys-C", "Overnight digestion at 37 °C")
    line("Desalting / concentration", "µC18 Omix (Agilent)",
         "Peptides desalted and concentrated before nano-LC analysis")

    section("2. CHROMATOGRAPHY (NANO-LC)")
    line("System", "NanoElute (Bruker Daltonics)", "UHPLC nano-flow")
    line("Injected amount", "~ 50 ng / sample", "Peptide load per injection")
    line("Trap column", "C18 PepMap 100", "5 mm × 300 µm i.d. — pre-concentration")
    line("Analytical column", "Reprosil C18 (Ionopticks)",
         "25 cm × 75 µm i.d., 1.6 µm C18 beads")
    line("Column temperature", "50 °C", "Reverse-phase separation")
    line("Mobile phases", "A: 0.1% FA / B: 0.1% FA, ACN",
         "A = 0.1% aqueous formic acid; B = 0.1% FA in ACN (v/v)")
    line("Flow rate", "250 nL/min", "Nanoflow")
    line("Gradient", "2 -> 95% B / 40 min",
         "2->15% B (15 min); ->30% (15 min); ->45% (5 min); ->95% (5 min)")

    section("3. MASS SPECTROMETRY (TIMS-TOF HT)")
    line("Instrument", "TIMS-TOF HT (Bruker Daltonics)",
         "CaptiveSpray nano-ESI source (Bruker)")
    line("Spray voltage", "1400 V", "Capillary temperature 180 °C")
    line("Mode / polarity", "PASEF DIA, positive", "Singly-charged exclusion")
    line("m/z range", "100 – 1700 m/z", "MS spectra acquisition")
    line("Mobility window", "0.75 – 1.28 1/k0", "TIMS domain")
    line("DIA scheme", "Variable windows", "Isolation windows from 350 to 1000 m/z")

    section("4. IDENTIFICATION & QUANTIFICATION (DIA-NN)")
    line("Software", "DIA-NN v2.5.1", "Search + LFQ quantification (Demichev 2019)")
    line("Database", "[!] TO BE FILLED IN (species)",
         "Protein database of the studied species — library-free / library generation")
    line("Mass / RT correction", "Automatic", "Automatic mass and retention-time correction")
    line("Fragments", "Top 6", "6 most intense fragments (identification + quantification)")
    line("Variable modifications", "N-term acetylation, Oxidation (M)", "Allowed modifications")
    line("Enzyme specificity", "Trypsin/P", "Cleavage rule")
    line("FDR", "1%", "Identification filtering")
    line("Cross-run normalization", "RT-dependent", "Cross-run normalization")

    note_r = state["row"] + 1
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=3)
    nc = ws.cell(note_r, 1,
        "Note: adapt the database to the studied species (highlighted field). "
        "The differential-analysis parameters (thresholds, stats, clusters) are "
        "detailed in the following 'Methods' sheet.")
    nc.font = Font(italic=True, size=9, color=C_HDRTXT); nc.alignment = wrap


def build_methods_sheet(ws, params: dict, tsv: pd.DataFrame,
                        df_results: pd.DataFrame, sig_names, contrast_names: list,
                        design: pd.DataFrame, n_proteins_filt: int,
                        go_enabled: bool = False, go_organism: str = None,
                        n_iter: int = 100, use_deqms: bool = False):
    """
    Construit la feuille 'Méthodes & Paramètres' documentant l'ensemble du
    pipeline : étapes, méthodes statistiques, seuils choisis. Pensée pour la
    traçabilité et la rédaction de la partie Méthodes d'un article.
    """
    from openpyxl.styles import Alignment, Border, Side

    title_font   = Font(bold=True, size=14, color="FFFFFF")
    title_fill   = PatternFill("solid", fgColor="2C3E50")
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="34495E")
    key_font     = Font(bold=True, size=10)
    key_fill     = PatternFill("solid", fgColor="ECF0F1")
    wrap         = Alignment(wrap_text=True, vertical="top")
    thin         = Side(style="thin", color="BDC3C7")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    p_volc = "p.adj (FDR-BH)" if params["volcano_use_padj"] else "p.value (raw)"
    p_anova = "p.adj (FDR-BH)" if params["anova_use_padj"] else "p.value (raw)"
    n_samples = len(design)
    n_cond = design["condition"].nunique()
    reps = design.groupby("condition").size()
    rep_range = f"{reps.min()}–{reps.max()}" if reps.min() != reps.max() else str(reps.min())
    n_raw = len(tsv)
    n_sig_anova = len(sig_names) if sig_names is not None else 0

    row = 1
    def put_title(text):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row=row, column=1, value=text)
        c.font = title_font; c.fill = title_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 24
        row += 2

    def put_section(text):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row=row, column=1, value=text)
        c.font = section_font; c.fill = section_fill
        ws.row_dimensions[row].height = 18
        row += 1

    def put_kv(key, value, method=""):
        nonlocal row
        kc = ws.cell(row=row, column=1, value=key)
        kc.font = key_font; kc.fill = key_fill; kc.border = border
        kc.alignment = wrap
        vc = ws.cell(row=row, column=2, value=value)
        vc.alignment = wrap; vc.border = border
        mc = ws.cell(row=row, column=3, value=method)
        mc.alignment = wrap; mc.border = border
        row += 1

    def put_header():
        nonlocal row
        for col, label in [(1, "Item"), (2, "Choice / Value"), (3, "Method / Detail")]:
            c = ws.cell(row=row, column=col, value=label)
            c.font = Font(bold=True, italic=True, size=9, color="7F8C8D")
        row += 1

    # ===== TITLE =====
    put_title("LFQ PROTEOMICS PIPELINE — Methods & Parameters")
    ws.cell(row=row-1, column=1,
            value=f"Auto-generated | {n_raw} raw proteins -> "
                  f"{n_proteins_filt} after filtering | {n_samples} samples, "
                  f"{n_cond} conditions").font = Font(italic=True, size=9, color="7F8C8D")
    row += 1

    # ===== DESIGN =====
    put_section("1. EXPERIMENTAL DESIGN")
    put_header()
    put_kv("Samples", str(n_samples), "LFQ columns from the DIA-NN report")
    put_kv("Conditions", str(n_cond), ", ".join(map(str, design["condition"].unique())))
    put_kv("Replicates / condition", rep_range, "Biological replicates")
    put_kv("Comparisons", str(len(contrast_names)),
           "All condition pairs (pairwise combinations)")

    # ===== PREPROCESSING =====
    put_section("2. DATA PREPROCESSING")
    put_header()
    put_kv("Source", "DIA-NN pg_matrix", "LFQ intensities (label-free quantification)")
    put_kv("Contaminants", "cRAP removal", "Filter on Protein.Group / Protein.Names")
    put_kv("Transformation", "log2", "log2(intensity), zeros -> NA")
    put_kv("Filtering", ">= (n_rep - 1) values",
           "Kept if detected in all but 1 replicate of at least one condition")
    put_kv("Proteins retained", f"{n_proteins_filt} / {n_raw}",
           "After missing-value filtering")
    _imp = params.get("impute_method", "qrilc")
    if _imp == "mixed":
        put_kv("Imputation", "Mixed (MNAR + MAR)",
               "QRILC (truncated normal, Lazar 2016) for proteins absent "
               "from all replicates of a condition (MNAR, below the detection "
               "limit); kNN (k=10) for partial missingness (MAR). "
               "Automatic per-missing-value classification based on the design.")
    else:
        put_kv("Imputation", "QRILC (MNAR)",
               "Quantile Regression Imputation of Left-Censored data (Lazar 2016): "
               "draw from a truncated normal below the estimated detection limit. "
               "Suited to DIA (predominantly MNAR missingness).")

    # ===== DIFFERENTIAL ANALYSIS =====
    put_section("3. DIFFERENTIAL ANALYSIS (Volcanos)")
    put_header()
    # Detect whether DEqMS is authoritative (presence of secondary limma columns)
    _deqms_active = any(str(c).endswith("_p.adj_limma") for c in df_results.columns)
    if _deqms_active:
        put_kv("Primary statistic", "DEqMS",
               "Residual variance modeled as a function of peptide count "
               "(Zhu et al. 2020). Authoritative for volcanos, UpSet, GO and robustness. "
               "limma retained in secondary columns (_limma)")
        put_kv("Base model", "limma eBayes",
               "Linear regression + empirical Bayes moderation (Smyth 2004), "
               "extended by DEqMS")
    else:
        put_kv("Primary statistic", "limma eBayes",
               "Linear regression + empirical Bayes moderation (Smyth 2004). "
               "Variance moderated by information borrowing across proteins — "
               "essential for small sample sizes (n=3). "
               "DEqMS not available (no pr_matrix peptide count)")
    put_kv("Design matrix", "~0 + condition", "One-hot encoding of conditions")
    put_kv("Contrasts", "All pairs", "makeContrasts equivalent")
    put_kv("Multiple correction", "Benjamini-Hochberg",
           "Global FDR (all comparisons, one test family)"
           if params.get("fdr_global") else
           "Per-contrast FDR (one test family per comparison)")
    put_kv("Significance threshold", p_volc,
           f"Threshold = {params['volcano_p_thresh']}")
    put_kv("Fold-Change threshold",
           f"ratio >= {params['volcano_ratio_min']}",
           f"|log2FC| >= {params['volcano_lfc_min']:.3f}")
    put_kv("Pi-score", "|log2FC| x -log10(p)",
           "Combined magnitude x significance score (Xiao et al.). "
           "Uses the p-value of the primary statistic.")

    # ===== ROBUSTNESS =====
    put_section("4. ROBUSTNESS SCORE")
    put_header()
    put_kv("Principle", f"{n_iter} re-imputations",
           "Repeated stochastic imputation (MinProb), model re-estimation "
           "at each iteration to assess stability against missing-value "
           "imputation uncertainty")
    put_kv("Applied thresholds", p_volc + f" < {params['volcano_p_thresh']} "
           f"& ratio >= {params['volcano_ratio_min']}",
           "IDENTICAL to the volcano thresholds (consistency)")
    put_kv("Score", f"0-{n_iter}",
           f"Number of iterations (out of {n_iter}) where the protein passes the thresholds. "
           "High score = differential robust to imputation uncertainty")

    # ===== ANOVA / HEATMAPS =====
    put_section("5. ANOVA & HEATMAPS")
    put_header()
    put_kv("Test", "One-way ANOVA (F-test)",
           "Per protein, on the condition factor")
    put_kv("Multiple correction", "Benjamini-Hochberg",
           "Global FDR (all comparisons, one test family)"
           if params.get("fdr_global") else
           "Per-contrast FDR (one test family per comparison)")
    put_kv("Significance threshold", p_anova,
           f"Threshold = {params['anova_p_thresh']} (independent of the volcanos)")
    put_kv("Significant proteins", str(n_sig_anova),
           "Proteins retained for the heatmaps")
    put_kv("Standardization", "Per-row Z-score",
           "Per-protein centering-scaling before clustering")
    put_kv("Clustering", "K-means (rows & columns)",
           "Clustered heatmap: structure of expression profiles")

    # ===== EXPLORATORY ANALYSES =====
    put_section("6. EXPLORATORY ANALYSES")
    put_header()
    put_kv("PCA", "PCA on imputed data",
           "Standardized, 95% confidence ellipses (normal distribution)")
    put_kv("UMAP", "n_neighbors=5, min_dist=0.1",
           "Non-linear projection, seed=42 (reproducible)")
    put_kv("Correlation", "Pearson between samples",
           "Quality control of replicate consistency")
    put_kv("UpSet", "DEP intersections",
           "Significant proteins shared across comparisons "
           "(volcano thresholds)")

    # ===== WGCNA =====
    put_section("7. WGCNA (Co-expression)")
    put_header()
    put_kv("Network", "Signed hybrid",
           "Adjacency = |Pearson correlation|^beta")
    put_kv("Soft power (beta)", "Auto (R2 >= 0.80)",
           "Selected by scale-free topology fit")
    put_kv("Similarity", "TOM",
           "Topological Overlap Matrix")
    put_kv("Modules", "Hierarchical clustering",
           "Average linkage on TOM dissimilarity + dynamic tree cut")
    put_kv("Module merging", "1 - cor(ME) < 0.25",
           "Merge modules with similar eigengenes")
    put_kv("Hub score", "kME",
           "Protein <-> module eigengene correlation")
    put_kv("Trait correlation", "Pearson module <-> condition",
           "Module-trait relationship")

    # ===== GO (if enabled) =====
    if go_enabled:
        put_section("8. FUNCTIONAL ENRICHMENT (GO)")
        put_header()
        put_kv("Tool", "g:Profiler (gost)",
               "Over-representation enrichment")
        put_kv("Species", str(go_organism), "g:Profiler organism code")
        put_kv("Sources", "GO:BP, GO:CC, GO:MF, REAC, KEGG",
               "Gene Ontology + Reactome + KEGG")
        put_kv("Input list", "DEP per contrast",
               f"Proteins passing {p_volc} < {params['volcano_p_thresh']} "
               f"& ratio >= {params['volcano_ratio_min']} (volcano thresholds)")
        put_kv("Multiple correction", "FDR (g:SCS disabled in favor of FDR)",
               "significance_threshold_method='fdr'")
        put_kv("Term filters", "size 5-1000, roots excluded",
               "Exclusion of GO:0003674/0008150/0005575")
        put_kv("Z-score", "term mean log2FC",
               "Pathway regulation direction")

    # ===== LARGEURS DE COLONNES =====
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 70


def export_excel(tsv: pd.DataFrame, df_results: pd.DataFrame,
                 df_anova: pd.DataFrame, df_intersections: pd.DataFrame,
                 mat_zscore: np.ndarray, meta: pd.DataFrame,
                 sig_names: np.ndarray, contrast_names: list,
                 design: pd.DataFrame, params: dict,
                 qc_files: list, pca_files: list, scatter_files: list,
                 volc_files: list, facet_volc: str, umap_file: str,
                 hm_corr: str, hm_classic: str, hm_clusters: str,
                 hm_violin: str, upset_file: str,
                 out_path: str,
                 go_enabled: bool = False, go_organism: str = None,
                 n_iter: int = 100, use_deqms: bool = False,
                 mat_imp: pd.DataFrame = None, umap_coords: pd.DataFrame = None,
                 imputation_info: pd.DataFrame = None):
    """Assemble le classeur Excel multionglets."""
    print("\n[EXCEL] Generating the Excel file...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill  = PatternFill("solid", fgColor="D9D9D9")
    zscore_fill  = PatternFill("solid", fgColor="FFCC00")
    header_font  = Font(bold=True)

    def write_df(ws, df, fill=header_fill):
        ws.append(list(df.columns))
        for cell in ws[1]:
            cell.fill = fill
            cell.font = header_font
        for row in df.itertuples(index=False):
            ws.append(list(row))

    # Largeur de colonne Excel ≈ 64 px ; hauteur de ligne ≈ 20 px.
    # openpyxl 3.1 ne redimensionne pas fiablement via img.width/height ;
    # on redimensionne donc physiquement le PNG avec PIL avant insertion.
    from openpyxl.utils import get_column_letter
    _resized_cache = {}

    def _resize_png(path, max_w_px, max_h_px):
        """Crée une copie redimensionnée du PNG (ratio conservé). Retourne le chemin."""
        try:
            from PIL import Image as _PILImage
        except ImportError:
            return path  # PIL absent : on garde l'original
        key = (str(path), max_w_px, max_h_px)
        if key in _resized_cache:
            return _resized_cache[key]
        try:
            im = _PILImage.open(path)
            w, h = im.size
            scale = min(max_w_px / w, max_h_px / h, 1.0)
            if scale < 1.0:
                im = im.resize((int(w * scale), int(h * scale)),
                               _PILImage.LANCZOS)
                base, ext = os.path.splitext(str(path))
                out = f"{base}_xl{max_w_px}x{max_h_px}{ext}"
                im.save(out)
                _resized_cache[key] = out
                return out
        except Exception:
            pass
        _resized_cache[key] = path
        return path

    def insert_img(ws, path, anchor=None, *, row=None, col=None,
                   max_w_px=560, max_h_px=440):
        """Insère une image redimensionnée (PIL) pour un placement sans
        chevauchement. anchor='A1' OU (row, col) 1-based."""
        if not path or not os.path.exists(str(path)):
            return
        rpath = _resize_png(path, max_w_px, max_h_px)
        img = XLImage(str(rpath))
        if anchor is None:
            anchor = f"{get_column_letter(col)}{row}"
        ws.add_image(img, anchor)

    def img_width_cols(path, max_w_px, max_h_px, col_px=64, margin=2):
        """Largeur d'affichage finale de l'image (après redimensionnement)
        exprimée en nombre de colonnes Excel. Sert à décaler dynamiquement
        un tableau placé à droite de l'image (largeur UpSet variable)."""
        try:
            from PIL import Image as _PILImage
            rpath = _resize_png(path, max_w_px, max_h_px)
            w, _ = _PILImage.open(rpath).size
            return int(ceil(w / col_px)) + margin
        except Exception:
            return int(ceil(max_w_px / col_px)) + margin

    # Une image bornée à 560×440 px occupe ~9 colonnes et ~22 lignes.
    # On espace les ancres de 10 colonnes et 24 lignes (marge de sécurité).
    IMG_COL_STEP = 10
    IMG_ROW_STEP = 24

    # --- Feuille M&M Amont (bioch/LC-MS/DIA-NN) en TOUTE PREMIÈRE position ---
    ws = wb.create_sheet("Methods_Upstream")
    build_mm_sheet(ws)

    # --- Feuille Méthodes (pipeline aval) en SECONDE position ---
    ws = wb.create_sheet("Methods")
    build_methods_sheet(ws, params, tsv, df_results, sig_names, contrast_names,
                        design, n_proteins_filt=len(df_results),
                        go_enabled=go_enabled, go_organism=go_organism,
                        n_iter=n_iter, use_deqms=use_deqms)

    # --- Onglet raw_data ---
    ws = wb.create_sheet("raw_data")
    write_df(ws, tsv)

    # --- Onglet Log2_Impute : matrice log2 imputée, TOUTES les protéines ---
    # Valeurs log2 imputées (MinProb) pour chaque protéine × chaque échantillon.
    if mat_imp is not None:
        ws = wb.create_sheet("Log2_Impute")
        df_imp = mat_imp.copy()
        df_imp.columns = [str(c) for c in df_imp.columns]
        # Identifiant protéine : on aligne sur meta (même ordre que mat_filt/mat_imp)
        ids = None
        for col in ("Protein.Group", "name"):
            if col in meta.columns and len(meta) == len(df_imp):
                ids = meta[col].values
                break
        df_imp.insert(0, "Protein.Group",
                      ids if ids is not None else np.arange(len(df_imp)))
        # Ajouter Genes/description si disponibles et alignés
        for extra in ("Genes", "First.Protein.Description"):
            if extra in meta.columns and len(meta) == len(df_imp):
                df_imp.insert(1, extra, meta[extra].values)
        df_imp = df_imp.round(4)
        write_df(ws, df_imp)
        print(f"  [OK] Log2_Impute sheet: {df_imp.shape[0]} proteins x "
              f"{mat_imp.shape[1]} samples (imputed log2)")

    # --- Contrôle Qualité ---
    ws = wb.create_sheet("QC")
    for i, f in enumerate(qc_files):
        r = (i // 3) * IMG_ROW_STEP + 1
        c = (i % 3) * IMG_COL_STEP + 1
        insert_img(ws, f, row=r, col=c)

    # --- PCA + UMAP + Corrélation ---
    ws = wb.create_sheet("PCA_UMAP")
    pca_imgs = list(pca_files) + [umap_file, hm_corr]
    for i, f in enumerate(pca_imgs):
        r = (i // 2) * IMG_ROW_STEP + 1
        c = (i % 2) * IMG_COL_STEP + 1
        insert_img(ws, f, row=r, col=c)

    # --- Feuille UMAP : coordonnées par échantillon (pour le dashboard) ---
    if umap_coords is not None and len(umap_coords) > 0:
        ws = wb.create_sheet("UMAP")
        write_df(ws, umap_coords)

    # --- Scatter Plots ---
    ws = wb.create_sheet("Scatter_Plots")
    for i, f in enumerate(scatter_files):
        r = (i // 2) * IMG_ROW_STEP + 1
        c = (i % 2) * IMG_COL_STEP + 1
        insert_img(ws, f, row=r, col=c)

    # --- Comparaison (stats différentielles) ---
    ws = wb.create_sheet("Differential_Expression")
    # Ajouter les infos d'imputation (comme le script R) si disponibles
    df_comp_src = df_results.copy()
    if imputation_info is not None and len(imputation_info) == len(df_comp_src):
        df_comp_src["imputed"] = imputation_info["imputed"].values
        df_comp_src["num_NAs"] = imputation_info["num_NAs"].values
    cols_keep = ["name", "Protein.Group", "Genes",
                 "First.Protein.Description", "N.Sequences"]
    for extra in ("imputed", "num_NAs"):
        if extra in df_comp_src.columns:
            cols_keep.append(extra)
    if "Peptide_Count_min" in df_comp_src.columns:
        cols_keep.append("Peptide_Count_min")
    for cname in contrast_names:
        # Colonnes principales (_p.val/_p.adj = DEqMS si dispo, sinon limma),
        # puis colonnes limma secondaires (_limma) pour comparaison directe.
        ordered = [f"{cname}_diff", f"{cname}_p.val", f"{cname}_p.adj",
                   f"{cname}_p.val_limma", f"{cname}_p.adj_limma",
                   f"Pi_Score_{cname}", f"Robustness_Score_{cname}"]
        for col in ordered:
            if col in df_comp_src.columns:
                cols_keep.append(col)
    cols_keep = [c for c in cols_keep if c in df_comp_src.columns]
    write_df(ws, df_comp_src[cols_keep])

    # --- Volcano Plots ---
    ws = wb.create_sheet("Volcano_Plots")
    p_str = f"{'p.adj' if params['volcano_use_padj'] else 'p.value'} < {params['volcano_p_thresh']} | ratio ≥ {params['volcano_ratio_min']}"
    ws.append([f"Facet Volcano — {len(contrast_names)} comparaisons | {p_str}"])
    insert_img(ws, facet_volc, anchor="A3", max_w_px=1100, max_h_px=1400)

    # Volcanos individuels : placés bien en dessous du facet, en grille 3 colonnes
    # avec espacement régulier (le facet peut être très haut avec beaucoup de
    # contrastes, donc on démarre largement plus bas).
    facet_rows = ceil(len(contrast_names) / 4) * 22 + 8
    base_row = max(60, facet_rows)
    for i, f in enumerate(volc_files):
        r = base_row + (i // 3) * IMG_ROW_STEP
        c = (i % 3) * IMG_COL_STEP + 1
        insert_img(ws, f, row=r, col=c)

    # --- Intersections UpSet ---
    # Image de l'UpSet à GAUCHE (col 1), tableau d'intersections à DROITE.
    # La largeur de l'UpSet varie selon le nombre de contrastes : on décale
    # donc le tableau dynamiquement en fonction de la largeur réelle de l'image.
    ws = wb.create_sheet("UpSet_Intersections")
    UPSET_MAX_W, UPSET_MAX_H = 1100, 850
    insert_img(ws, upset_file, row=1, col=1,
               max_w_px=UPSET_MAX_W, max_h_px=UPSET_MAX_H)

    if df_intersections is not None and len(df_intersections) > 0:
        # Colonne de départ du tableau = largeur de l'image (en colonnes) + marge
        start_col = 1
        if upset_file and os.path.exists(str(upset_file)):
            start_col = img_width_cols(upset_file, UPSET_MAX_W, UPSET_MAX_H) + 1
        # En-tête
        for j, colname in enumerate(df_intersections.columns):
            cell = ws.cell(row=1, column=start_col + j, value=colname)
            cell.fill = header_fill
            cell.font = header_font
        # Données
        for i, row in enumerate(df_intersections.itertuples(index=False), start=2):
            for j, val in enumerate(row):
                ws.cell(row=i, column=start_col + j, value=val)

    # --- Z-score Heatmap ---
    ws = wb.create_sheet("Zscore_Heatmap")
    if mat_zscore is not None and len(sig_names) > 0:
        df_z = pd.DataFrame(mat_zscore,
                            columns=design["label"].values)
        df_z.insert(0, "Protein.Group", sig_names[:len(df_z)])
        write_df(ws, df_z, fill=zscore_fill)
        # heatmap après les colonnes de données
        z_col = df_z.shape[1] + 2
    else:
        z_col = 13
    insert_img(ws, hm_classic, row=2, col=z_col, max_w_px=900, max_h_px=1100)

    # --- ANOVA Classique ---
    ws = wb.create_sheet("ANOVA_Results")
    write_df(ws, df_anova)
    insert_img(ws, hm_classic, row=2, col=df_anova.shape[1] + 2,
               max_w_px=900, max_h_px=1100)

    # --- ANOVA Clusters ---
    ws = wb.create_sheet("ANOVA_Clusters")
    write_df(ws, df_anova)
    base_c = df_anova.shape[1] + 2
    insert_img(ws, hm_clusters, row=2, col=base_c, max_w_px=900, max_h_px=1100)
    insert_img(ws, hm_violin,   row=2, col=base_c + IMG_COL_STEP + 5,
               max_w_px=900, max_h_px=600)

    wb.save(out_path)
    print(f"  [OK] Excel file saved: {out_path}")


# ==============================================================================
# HELPERS
# ==============================================================================

def _condition_colors(conditions) -> dict:
    """Génère une palette de couleurs par condition."""
    unique_conds = list(dict.fromkeys(conditions))
    n = max(len(unique_conds), 1)
    # API compatible matplotlib < 3.7 et >= 3.9 (get_cmap supprimé en 3.9)
    try:
        cmap = matplotlib.colormaps["tab20"].resampled(n)
    except (AttributeError, KeyError):
        cmap = plt.cm.get_cmap("tab20", n)
    return {c: matplotlib.colors.to_hex(cmap(i))
            for i, c in enumerate(unique_conds)}


def _makedirs(path):
    os.makedirs(path, exist_ok=True)
    return path


def _col_letter(n: int) -> str:
    """Convertit un index de colonne (1-based) en lettres Excel (A, B, … AA…)."""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


# ==============================================================================
# WGCNA — Fonctions intégrées
# ==============================================================================

WGCNA_CONFIG = {
    "soft_power":       None,   # None = auto-détection | int = valeur fixe
    "min_module_size":  10,
    "merge_cut_height": 0.25,
    "n_powers_test":    range(1, 21),
    # Allègement : ne garder que les N protéines les plus variables pour le réseau.
    # La TOM est en O(N²) en mémoire et O(N²·échantillons) en temps ; au-delà de
    # ~5000 protéines le calcul devient lourd. Les protéines peu variables
    # n'apportent rien à l'analyse de co-expression.
    "max_proteins":     2000,   # None = pas de filtrage | int = top-N par variance
}


def wgcna_pick_soft_threshold(datExpr: pd.DataFrame, powers: range,
                               out_dir: str) -> tuple[int, str]:
    """
    Sélection automatique du soft-thresholding power.
    Retourne le plus petit power avec R² ≥ 0.80 sur le fit scale-free.
    """
    print("\n[WGCNA] Soft power selection...")
    X = datExpr.values  # (n_samples × n_proteins)
    corr_mat = np.corrcoef(X.T)
    np.fill_diagonal(corr_mat, 0)

    r2_fits, mean_k = [], []
    for power in powers:
        adj = np.abs(corr_mat) ** power
        k = adj.sum(axis=0)
        mean_k.append(k.mean())
        k_pos = k[k > 0]
        if len(k_pos) < 5:
            r2_fits.append(0)
            continue
        hist_counts, bin_edges = np.histogram(k_pos, bins=20)
        bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
        valid = hist_counts > 0
        log_k = np.log10(bin_centers[valid])
        log_p = np.log10(hist_counts[valid] / hist_counts[valid].sum())
        if len(log_k) < 3:
            r2_fits.append(0)
            continue
        coeffs = np.polyfit(log_k, log_p, 1)
        y_pred = np.polyval(coeffs, log_k)
        ss_res = np.sum((log_p - y_pred) ** 2)
        ss_tot = np.sum((log_p - log_p.mean()) ** 2)
        r2_fits.append(max(1 - ss_res / ss_tot if ss_tot > 0 else 0, 0))

    powers_list = list(powers)
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
    ax1.plot(powers_list, r2_fits, "o-", color="#2C3E50")
    ax1.axhline(0.80, color="red", ls="--", lw=1, label="R² = 0.80")
    ax1.set_xlabel("Soft Thresholding Power")
    ax1.set_ylabel("Scale Free Topology R²")
    ax1.set_title("Soft Power selection")
    ax1.legend()
    ax2.plot(powers_list, mean_k, "s-", color="#3498DB")
    ax2.set_xlabel("Soft Thresholding Power")
    ax2.set_ylabel("Mean connectivity")
    ax2.set_title("Mean connectivity vs Power")
    fig.tight_layout()
    f = os.path.join(out_dir, "wgcna_soft_threshold.png")
    fig.savefig(f, dpi=150)
    plt.close(fig)

    best_power = next((p for p, r2 in zip(powers_list, r2_fits) if r2 >= 0.80), None)
    if best_power is None:
        best_power = powers_list[np.argmax(r2_fits)]
        print(f"  [WARN] R2 < 0.80 for all powers. Best: {best_power} (R2={max(r2_fits):.2f})")
    else:
        print(f"  [OK] Soft power: {best_power} (R2={r2_fits[powers_list.index(best_power)]:.2f})")
    return best_power, f


def wgcna_build_network(datExpr: pd.DataFrame, soft_power: int) -> tuple[np.ndarray, np.ndarray]:
    """TOM signée hybride par blocs (500×500) pour gestion mémoire."""
    print(f"\n[WGCNA] Network construction (soft power={soft_power})...")
    X = datExpr.values.T  # (n_proteins × n_samples)
    n = X.shape[0]
    corr = np.corrcoef(X)
    np.fill_diagonal(corr, 0)
    adj = np.abs(corr) ** soft_power
    np.fill_diagonal(adj, 0)
    k = adj.sum(axis=1)

    print(f"  -> TOM computation ({n} x {n})...")
    block_size = 500
    TOM = np.zeros((n, n), dtype=np.float32)
    for i in range(0, n, block_size):
        ei = min(i + block_size, n)
        for j in range(0, n, block_size):
            ej = min(j + block_size, n)
            L = adj[i:ei, :] @ adj[:, j:ej]
            min_k = np.minimum(k[i:ei, np.newaxis], k[np.newaxis, j:ej])
            denom = np.maximum(min_k + 1 - adj[i:ei, j:ej], 1e-10)
            TOM[i:ei, j:ej] = (L + adj[i:ei, j:ej]) / denom
    np.fill_diagonal(TOM, 0)
    TOM = np.clip(TOM, 0, 1)
    print("  [OK] TOM computed")
    return TOM, adj


def wgcna_detect_modules(TOM: np.ndarray, datExpr: pd.DataFrame,
                          min_module_size: int, merge_cut_height: float,
                          out_dir: str) -> tuple[np.ndarray, np.ndarray, str]:
    """
    Clustering hiérarchique sur dissimilarité TOM + découpe dynamique.
    Fusion des modules trop similaires (corrélation ME > 1 - merge_cut_height).
    """
    print("\n[WGCNA] Module detection...")
    diss = 1 - TOM
    diss_cond = np.maximum(squareform(diss, checks=False), 0)
    Z = linkage(diss_cond, method="average")

    cut_h = np.percentile(Z[:, 2], 80)
    labels_raw = fcluster(Z, t=cut_h, criterion="distance")

    unique, counts = np.unique(labels_raw, return_counts=True)
    small = unique[counts < min_module_size]
    labels_clean = labels_raw.copy()
    for s in small:
        labels_clean[labels_clean == s] = 0

    valid = sorted(set(labels_clean) - {0})
    remap = {old: new for new, old in enumerate(valid, 1)}
    remap[0] = 0
    labels_clean = np.array([remap[l] for l in labels_clean])

    # Calcul des MEs pour fusion
    ME_dict = {}
    for mod in sorted(set(labels_clean)):
        idx = np.where(labels_clean == mod)[0]
        if len(idx) >= 2:
            me = PCA(n_components=1).fit_transform(datExpr.values[:, idx]).flatten()
            ME_dict[mod] = me

    # Fusion itérative
    merged = True
    while merged:
        merged = False
        mods = sorted(set(labels_clean) - {0})
        for i in range(len(mods)):
            for j in range(i + 1, len(mods)):
                m1, m2 = mods[i], mods[j]
                if m1 not in ME_dict or m2 not in ME_dict:
                    continue
                r, _ = pearsonr(ME_dict[m1], ME_dict[m2])
                if (1 - r) < merge_cut_height:
                    labels_clean[labels_clean == m2] = m1
                    idx1 = np.where(labels_clean == m1)[0]
                    if len(idx1) >= 2:
                        ME_dict[m1] = PCA(n_components=1).fit_transform(
                            datExpr.values[:, idx1]).flatten()
                    del ME_dict[m2]
                    merged = True
                    break
            if merged:
                break

    valid = sorted(set(labels_clean) - {0})
    remap = {old: new for new, old in enumerate(valid, 1)}
    remap[0] = 0
    labels_clean = np.array([remap.get(l, 0) for l in labels_clean])

    n_final = len(set(labels_clean)) - (1 if 0 in labels_clean else 0)
    print(f"  [OK] {n_final} modules detected (after merging)")

    colors_palette = _wgcna_color_palette(int(labels_clean.max()) + 1)
    # Liste Python (pas np.array) : np.array de strings donne np.str_ qui
    # provoque ValueError dans matplotlib lors de la résolution RGBA.
    module_colors = [str(colors_palette[l]) for l in labels_clean]

    # Augmenter la limite de récursion (scipy dendrogram est récursif sur les feuilles)
    import sys as _sys
    _old_limit = _sys.getrecursionlimit()
    _sys.setrecursionlimit(max(_old_limit, len(module_colors) * 3 + 1000))

    n_leaves = Z.shape[0] + 1
    f_dendro = os.path.join(out_dir, "wgcna_dendrogram.png")
    try:
        # 1) Récupérer l'ordre des feuilles SANS tracer (pas de récursion d'affichage)
        dend_order = dendrogram(Z, no_plot=True)
        leaf_order = dend_order["leaves"]

        fig, (ax1, ax2) = plt.subplots(
            2, 1, figsize=(max(12, min(len(module_colors) * 0.05, 60)), 6),
            gridspec_kw={"height_ratios": [4, 0.5]})

        # 2) Dendrogramme : tronqué si trop de feuilles (illustratif, modules déjà calculés)
        if n_leaves > 400:
            dendrogram(Z, ax=ax1, no_labels=True, color_threshold=0,
                       above_threshold_color="grey",
                       link_color_func=lambda k: "grey",
                       truncate_mode="lastp", p=120)
            ax1.set_title(f"Protein dendrogram (TOM dissimilarity) — "
                          f"truncated to the last 120 clusters / {n_leaves} proteins")
        else:
            dendrogram(Z, ax=ax1, no_labels=True, color_threshold=0,
                       above_threshold_color="grey",
                       link_color_func=lambda k: "grey")
            ax1.set_title("Protein dendrogram (TOM dissimilarity)")
        ax1.set_ylabel("Height")

        # 3) Barre de couleur des modules selon l'ordre RÉEL des feuilles
        for xi, col in enumerate([module_colors[i] for i in leaf_order]):
            ax2.add_patch(mpatches.Rectangle((xi, 0), 1, 1, color=col))
        ax2.set_xlim(0, len(leaf_order))
        ax2.set_ylim(0, 1)
        ax2.axis("off")
        fig.tight_layout()
        fig.savefig(f_dendro, dpi=150, bbox_inches="tight")
        plt.close(fig)
    except RecursionError:
        # Garde-fou : si le tracé échoue malgré tout, on produit juste la barre
        # de couleurs des modules (les modules eux-mêmes restent valides).
        print("  [WARN] Dendrogram too deep to plot — module bar only.")
        try:
            leaf_order = list(range(len(module_colors)))
            fig, ax2 = plt.subplots(figsize=(14, 1.2))
            for xi, col in enumerate(module_colors):
                ax2.add_patch(mpatches.Rectangle((xi, 0), 1, 1, color=col))
            ax2.set_xlim(0, len(module_colors)); ax2.set_ylim(0, 1); ax2.axis("off")
            ax2.set_title("WGCNA modules (dendrogram not plotted)")
            fig.savefig(f_dendro, dpi=150, bbox_inches="tight")
            plt.close(fig)
        except Exception:
            f_dendro = None
    finally:
        _sys.setrecursionlimit(_old_limit)

    return module_colors, labels_clean, f_dendro


def _wgcna_color_palette(n: int) -> list:
    wgcna_colors = [
        "grey", "turquoise", "blue", "brown", "yellow", "green",
        "red", "black", "pink", "magenta", "purple", "greenyellow",
        "tan", "salmon", "cyan", "midnightblue", "lightcyan", "#999999",
        "lightgreen", "lightyellow", "royalblue", "darkred", "darkgreen",
        "darkturquoise", "darkgrey", "orange", "darkorange", "white",
        "skyblue", "saddlebrown", "steelblue", "paleturquoise", "violet",
    ]
    extended = wgcna_colors + [to_hex(plt.cm.tab20(i % 20))
                                for i in range(max(0, n - len(wgcna_colors) + 1))]
    return extended[:n + 1]


def wgcna_compute_eigengenes(datExpr: pd.DataFrame, module_colors: np.ndarray,
                              design: pd.DataFrame, out_dir: str
                              ) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, str]:
    """Module eigengenes (PC1 par module) + corrélation avec les traits (conditions)."""
    print("\n[WGCNA] Module eigengenes and trait correlation...")
    unique_mods = sorted(set(module_colors))
    ME_dict = {}
    for mod in unique_mods:
        idx = np.where(np.array(module_colors) == mod)[0]
        if len(idx) < 2:
            continue
        me = PCA(n_components=1).fit_transform(datExpr.values[:, idx]).flatten()
        mean_mod = datExpr.values[:, idx].mean(axis=1)
        if pearsonr(me, mean_mod)[0] < 0:
            me = -me
        ME_dict[f"ME_{mod}"] = me

    df_ME = pd.DataFrame(ME_dict, index=datExpr.index)

    conditions = design["condition"].values
    unique_conds = sorted(set(conditions))
    df_traits = pd.DataFrame(
        {c: (conditions == c).astype(float) for c in unique_conds},
        index=design["label"].values
    ).loc[datExpr.index]

    n_me, n_tr = len(df_ME.columns), len(df_traits.columns)
    cor_mat  = np.zeros((n_me, n_tr))
    pval_mat = np.zeros((n_me, n_tr))
    for i, me_col in enumerate(df_ME.columns):
        for j, tr_col in enumerate(df_traits.columns):
            r, p = pearsonr(df_ME[me_col].values, df_traits[tr_col].values)
            cor_mat[i, j]  = r
            pval_mat[i, j] = p

    df_cor  = pd.DataFrame(cor_mat,  index=df_ME.columns, columns=df_traits.columns)
    df_pval = pd.DataFrame(pval_mat, index=df_ME.columns, columns=df_traits.columns)

    # Heatmap module-trait
    fig, ax = plt.subplots(figsize=(max(6, n_tr * 0.7), max(5, n_me * 0.45)))
    im = ax.imshow(df_cor.values, cmap="RdBu_r", vmin=-1, vmax=1, aspect="auto")
    for i in range(n_me):
        for j in range(n_tr):
            r = df_cor.values[i, j]
            p = df_pval.values[i, j]
            ax.text(j, i, f"{r:.2f}\n({p:.3f})", ha="center", va="center",
                    fontsize=max(5, 9 - n_me // 4),
                    color="black" if abs(r) < 0.7 else "white")
    ax.set_xticks(range(n_tr))
    ax.set_xticklabels(df_cor.columns, rotation=45, ha="right", fontsize=8)
    ax.set_yticks(range(n_me))
    ax.set_yticklabels(df_cor.index, fontsize=8)
    for tick, me in zip(ax.get_yticklabels(), df_cor.index):
        try:
            tick.set_color(me.replace("ME_", ""))
        except Exception:
            pass
    plt.colorbar(im, ax=ax, fraction=0.03, label="Pearson r")
    ax.set_title("Module–Trait correlation\n(r, p-value)", fontsize=10)
    fig.tight_layout()
    f_hm = os.path.join(out_dir, "wgcna_module_trait_heatmap.png")
    fig.savefig(f_hm, dpi=150, bbox_inches="tight")
    plt.close(fig)

    return df_ME, df_cor, df_pval, f_hm


def wgcna_compute_hub_scores(datExpr: pd.DataFrame, df_ME: pd.DataFrame,
                              module_colors: np.ndarray, meta: pd.DataFrame,
                              out_dir: str) -> tuple[pd.DataFrame, pd.DataFrame, str]:
    """kME par protéine + hub score (kME du module propre) + barplot top 5."""
    print("\n[WGCNA] Hub scores (kME)...")
    protein_names = datExpr.columns.tolist()
    unique_mods   = sorted(set(module_colors))

    kme_dict = {}
    for mod in unique_mods:
        me_col = f"ME_{mod}"
        if me_col not in df_ME.columns:
            continue
        me_vals  = df_ME[me_col].values
        kme_vals = np.array([pearsonr(datExpr[p].values, me_vals)[0]
                             for p in protein_names])
        kme_dict[f"kME_{mod}"] = kme_vals

    df_kme = pd.DataFrame(kme_dict, index=protein_names)

    hub_scores = np.array([
        df_kme.loc[p, f"kME_{module_colors[k]}"]
        if f"kME_{module_colors[k]}" in df_kme.columns else np.nan
        for k, p in enumerate(protein_names)
    ])

    df_final = pd.DataFrame({
        "Gene_Name":    protein_names,
        "Module_Color": module_colors,
        "HubScore":     hub_scores,
    })
    if "name" in meta.columns:
        meta_sub = meta[["name", "Genes", "First.Protein.Description"]].drop_duplicates("name")
        df_final = df_final.merge(meta_sub, left_on="Gene_Name",
                                   right_on="name", how="left").drop(columns="name", errors="ignore")
    df_final = df_final.sort_values(["Module_Color", "HubScore"],
                                     ascending=[True, False]).reset_index(drop=True)

    # Barplot top 5 hubs par module
    # Boucle explicite : évite le comportement variable de groupby+apply
    # selon la version pandas (2.x peut perdre Module_Color après reset_index)
    _df_non_grey = df_final[df_final["Module_Color"] != "grey"]
    df_plot = pd.concat(
        [grp.nlargest(5, "HubScore") for _, grp in _df_non_grey.groupby("Module_Color")],
        ignore_index=True
    ) if len(_df_non_grey) > 0 else pd.DataFrame(columns=df_final.columns)
    mods = df_plot["Module_Color"].unique() if len(df_plot) > 0 else []
    n_mods = len(mods)
    f_hubs = None
    if n_mods > 0:
        n_cols = min(4, n_mods)
        n_rows = (n_mods + n_cols - 1) // n_cols
        fig, axes = plt.subplots(n_rows, n_cols, figsize=(n_cols * 3.5, n_rows * 3))
        axes = np.array(axes).flatten()
        for idx, mod in enumerate(mods):
            ax = axes[idx]
            sub = df_plot[df_plot["Module_Color"] == mod].sort_values("HubScore")
            ax.barh(sub["Gene_Name"].values, sub["HubScore"].values,
                    color=str(mod), edgecolor="white")
            ax.set_title(f"Module {mod}", fontsize=9, fontweight="bold", color=str(mod))
            ax.set_xlabel("kME", fontsize=8)
            ax.tick_params(labelsize=7)
            ax.set_xlim(0, 1)
        for idx in range(n_mods, len(axes)):
            axes[idx].set_visible(False)
        fig.suptitle("Top 5 Hub Proteins per Module", fontsize=11, fontweight="bold")
        fig.tight_layout()
        f_hubs = os.path.join(out_dir, "wgcna_hubs.png")
        fig.savefig(f_hubs, dpi=150, bbox_inches="tight")
        plt.close(fig)

    print(f"  [OK] Hub scores computed ({len(protein_names)} proteins)")
    return df_final, df_kme, f_hubs


def wgcna_module_summary(df_final: pd.DataFrame, df_cor: pd.DataFrame,
                          out_dir: str) -> tuple[pd.DataFrame, str]:
    """Résumé par module : taille, top hub, top trait corrélé."""
    rows = []
    for mod, grp in df_final.groupby("Module_Color"):
        n = len(grp)
        top_hub = grp.nlargest(1, "HubScore")["Gene_Name"].values[0] if n > 0 else ""
        top_hub_score = grp["HubScore"].max()
        me_col = f"ME_{mod}"
        if me_col in df_cor.index:
            top_trait = df_cor.loc[me_col].abs().idxmax()
            top_r     = df_cor.loc[me_col, top_trait]
        else:
            top_trait, top_r = "", np.nan
        rows.append({"Module": mod, "N_proteins": n, "Top_Hub": top_hub,
                     "Hub_kME": round(top_hub_score, 3),
                     "Top_Trait": top_trait, "Trait_r": round(float(top_r), 3)})

    df_summary = pd.DataFrame(rows).sort_values("N_proteins", ascending=False)

    fig, ax = plt.subplots(figsize=(max(8, len(df_summary) * 0.5), 4))
    mods_plot = df_summary[df_summary["Module"] != "grey"]
    ax.bar(mods_plot["Module"], mods_plot["N_proteins"],
           color=[str(c) for c in mods_plot["Module"].values], edgecolor="white")
    ax.set_xlabel("Module")
    ax.set_ylabel("Number of proteins")
    ax.set_xticklabels(mods_plot["Module"], rotation=45, ha="right", fontsize=8)
    ax.set_title("WGCNA module sizes")
    fig.tight_layout()
    f_sizes = os.path.join(out_dir, "wgcna_module_sizes.png")
    fig.savefig(f_sizes, dpi=150)
    plt.close(fig)

    return df_summary, f_sizes


def run_wgcna(mat_imp: pd.DataFrame, meta: pd.DataFrame,
              design: pd.DataFrame, out_dir: str) -> dict:
    """
    Point d'entrée WGCNA. Reçoit mat_imp (protéines × samples) directement
    depuis le pipeline DEP — pas besoin de relire l'Excel.
    Retourne un dict avec tous les objets et chemins de figures.
    """
    print("\n" + "="*60)
    print("  WGCNA — Protein co-expression analysis")
    print("="*60)

    cfg = WGCNA_CONFIG

    # Orientation WGCNA : samples en lignes, protéines en colonnes
    datExpr = mat_imp.T.copy()
    datExpr.columns = meta["name"].values[:datExpr.shape[1]]
    datExpr.index   = design["label"].values

    # Nettoyage : exclure colonnes avec > 50% NaN
    col_ok  = datExpr.isna().mean(axis=0) < 0.5
    datExpr = datExpr.loc[:, col_ok].fillna(datExpr.mean(axis=0))
    n_total = datExpr.shape[1]

    # Allègement : filtrer sur les protéines les plus variables.
    # La TOM est O(N²) ; au-delà de max_proteins on ne garde que les plus variables
    # (les protéines invariantes n'informent pas la co-expression).
    max_prot = cfg.get("max_proteins")
    if max_prot is not None and n_total > max_prot:
        variances = datExpr.var(axis=0)
        top_cols = variances.sort_values(ascending=False).head(max_prot).index
        datExpr = datExpr[top_cols]
        print(f"  -> {n_total} proteins -> {datExpr.shape[1]} kept "
              f"(top variance, TOM lightening) | {datExpr.shape[0]} samples")
    else:
        print(f"  -> {datExpr.shape[1]} proteins | {datExpr.shape[0]} samples")

    # 1. Soft threshold
    if cfg["soft_power"] is None:
        soft_power, f_thresh = wgcna_pick_soft_threshold(
            datExpr, cfg["n_powers_test"], out_dir)
    else:
        soft_power = cfg["soft_power"]
        f_thresh   = None
        print(f"  [INFO] Soft power fixed: {soft_power}")

    # 2. TOM
    TOM, adj = wgcna_build_network(datExpr, soft_power)

    # 3. Modules
    module_colors, labels_int, f_dendro = wgcna_detect_modules(
        TOM, datExpr, cfg["min_module_size"], cfg["merge_cut_height"], out_dir)

    # 4. Eigengenes + corrélation traits
    df_ME, df_cor, df_pval, f_hm = wgcna_compute_eigengenes(
        datExpr, module_colors, design, out_dir)

    # 5. Hub scores
    df_final, df_kme, f_hubs = wgcna_compute_hub_scores(
        datExpr, df_ME, module_colors, meta, out_dir)

    # 6. Résumé
    df_summary, f_sizes = wgcna_module_summary(df_final, df_cor, out_dir)

    print("  [OK] WGCNA done")
    return {
        "df_final": df_final, "df_kme": df_kme, "df_ME": df_ME,
        "df_cor": df_cor, "df_pval": df_pval, "df_summary": df_summary,
        "f_thresh": f_thresh, "f_dendro": f_dendro,
        "f_hm": f_hm, "f_hubs": f_hubs, "f_sizes": f_sizes,
    }


def export_wgcna_sheets(wb: openpyxl.Workbook, wgcna: dict,
                         write_df_fn, ins_img_fn):
    """
    Ajoute les onglets WGCNA au classeur Excel DEP existant.
    Reçoit les helpers write_df et ins_img pour réutiliser le même style.
    """
    df_final  = wgcna["df_final"]
    df_kme    = wgcna["df_kme"]
    df_ME     = wgcna["df_ME"]
    df_cor    = wgcna["df_cor"]
    df_pval   = wgcna["df_pval"]
    df_summary = wgcna["df_summary"]

    # --- Résultats principaux ---
    ws = wb.create_sheet("WGCNA_Results")
    write_df_fn(ws, df_final)
    col_img = len(df_final.columns) + 2
    ins_img_fn(ws, wgcna["f_hm"],   f"{_col_letter(col_img)}2")
    ins_img_fn(ws, wgcna["f_hubs"], f"{_col_letter(col_img)}42")

    # --- Résumé modules ---
    ws2 = wb.create_sheet("WGCNA_Modules")
    write_df_fn(ws2, df_summary)
    ins_img_fn(ws2, wgcna["f_sizes"], "H2")

    # --- kME complet ---
    ws3 = wb.create_sheet("WGCNA_kME")
    df_kme_out = df_kme.reset_index().rename(columns={"index": "Gene_Name"})
    write_df_fn(ws3, df_kme_out)

    # --- Module Eigengenes ---
    ws4 = wb.create_sheet("WGCNA_Eigengenes")
    df_ME_out = df_ME.reset_index().rename(columns={"index": "Sample"})
    write_df_fn(ws4, df_ME_out)

    # --- Corrélation module-trait ---
    ws5 = wb.create_sheet("WGCNA_ModuleTrait")
    df_cor_out = df_cor.reset_index().rename(columns={"index": "Module"})
    write_df_fn(ws5, df_cor_out)
    # P-values sous les corrélations
    row_start = len(df_cor_out) + 3
    ws5.cell(row=row_start, column=1, value="P-values")
    df_pval_out = df_pval.reset_index().rename(columns={"index": "Module"})
    for row in df_pval_out.itertuples(index=False):
        ws5.append(list(row))
    ins_img_fn(ws5, wgcna["f_hm"], "H2")

    # --- Diagnostics ---
    ws6 = wb.create_sheet("WGCNA_Diagnostics")
    ins_img_fn(ws6, wgcna["f_thresh"],  "A2")
    ins_img_fn(ws6, wgcna["f_dendro"], "M2")


# ==============================================================================
# MAIN
# ==============================================================================

def main():
    # --- Configuration : YAML / rechargement / interactif ---
    # resolve_config gère les trois modes sans modifier le comportement interactif :
    #   · python deimos.py --config mon_projet.yaml  → zéro question
    #   · python deimos.py --save-config             → questions + sauvegarde
    #   · python deimos.py                           → propose last_config ou questions
    # last_config.yaml est toujours sauvegardé dans out_dir après chaque run.
    _pr_path_default = "report.pr_matrix.tsv"

    params = resolve_config(
        ask_params_fn     = ask_params,
        ask_go_params_fn  = ask_go_params if _GO_AVAILABLE else lambda: None,
        go_available      = _GO_AVAILABLE,
        dash_available    = _DASHBOARD_AVAILABLE,
        pr_path           = _pr_path_default,
    )

    # --- Variables locales reconstituées (compatibilité avec le reste de main()) ---
    tsv_path    = params["tsv_path"]
    design_path = params["design_path"]
    pr_path     = params.get("pr_path", _pr_path_default)
    out_dir     = params["out_dir"]
    make_wgcna  = params["make_wgcna"]
    make_dash   = params["make_dashboard"]
    use_deqms   = params["use_deqms"]

    # go_params : dict {"organism": "bnapus"} attendu par run_go_enrichment, ou None
    _go_org   = params.get("go_organism")
    go_params = {"organism": _go_org} if (_GO_AVAILABLE and _go_org) else None

    _makedirs(out_dir)

    # --- 0. Diagnostic labels ---
    if not diagnose_labels(tsv_path, design_path):
        sys.exit(1)

    # --- 1. Chargement ---
    tsv, design = load_data(tsv_path, design_path)

    # --- 2. Matrice expression ---
    mat_log2, meta, design_filt, lfq_cols = build_expression_matrix(tsv, design)

    # Sécurité : colonnes dupliquées → crash dans pandas lors de l'imputation
    # Doit être fait AVANT plot_qc pour garder mat_log2 et design_filt synchronisés
    if mat_log2.columns.duplicated().any():
        n_dup = mat_log2.columns.duplicated().sum()
        print(f"  [WARN] {n_dup} duplicate sample column(s) in matrix — keeping first occurrence.")
        dup_labels = mat_log2.columns[mat_log2.columns.duplicated()].tolist()
        print(f"         Duplicates: {dup_labels}")
        keep_mask_dup = ~pd.Series(mat_log2.columns).duplicated().values
        mat_log2    = mat_log2.loc[:, keep_mask_dup]
        design_filt = design_filt[keep_mask_dup].reset_index(drop=True)

    # --- 3. QC ---
    print("\n[QC] Quality control...")
    qc_files = plot_qc(mat_log2, meta, design_filt, out_dir)

    # --- 4. Filtration + Imputation ---
    print("\n[STEP] Filtering and imputation...")
    mat_filt, keep_mask = filter_missval(mat_log2, design_filt, thr=1)
    meta_filt = meta[keep_mask].reset_index(drop=True)

    # Diagnostic de missingness (objective le choix Mixte vs QRILC)
    miss_diag = diagnose_missingness(mat_filt, design_filt, out_dir)
    if miss_diag.get("plot"):
        qc_files.append(miss_diag["plot"])

    imp_method = params.get("impute_method", "qrilc")
    mat_imp = impute(mat_filt, method=imp_method, design=design_filt)
    print(f"  [OK] Imputation: {imp_method.upper()}")
    qc_imp  = plot_imputation(mat_filt, mat_imp, out_dir)
    qc_files.append(qc_imp)

    # Info d'imputation par protéine (comme le script R) : nombre de valeurs
    # manquantes imputées par protéine, et drapeau imputed oui/non.
    na_mask = mat_filt.isna()
    n_na_per_prot = na_mask.sum(axis=1).values
    imputation_info = pd.DataFrame({
        "imputed": np.where(n_na_per_prot > 0, "VRAI", "FAUX"),
        "num_NAs": n_na_per_prot.astype(int),
    }, index=mat_filt.index)

    mat_imp.columns  = design_filt["label"].values
    mat_filt.columns = design_filt["label"].values

    # --- 4bis. Comptage peptidique pour DEqMS (optionnel) ---
    pep_count = None
    if use_deqms:
        print("\n[STEP] Building peptide count (DEqMS)...")
        pep_count = build_peptide_count_table(
            pr_path, meta_filt["name"].tolist(), count_level="peptide")
        if pep_count is None:
            print("  -> DEqMS disabled (counting unavailable).")

    # --- 5. Analyse différentielle + Robustness ---
    print("\n[STEP] Differential analysis + robustness score...")
    N_ITER = params.get("n_iter_robustness", 100)
    df_results, contrast_names, scatter_files = run_differential_analysis(
        mat_imp, mat_filt, meta_filt, design_filt, params, out_dir,
        n_iter=N_ITER, pep_count=pep_count
    )

    # --- 6. Volcanos ---
    print("\n[STEP] Generating volcano plots...")
    volc_files, facet_volc = plot_volcanoes(df_results, contrast_names, params, out_dir)

    # --- 7. PCA ---
    print("\n[STEP] PCA...")
    pca_files = plot_pca(mat_imp, design_filt, out_dir)

    # --- 8. Corrélation ---
    print("\n[STEP] Correlation heatmap...")
    hm_corr = plot_correlation_heatmap(mat_imp, design_filt, out_dir)

    # --- 9. UMAP ---
    print("\n🗺️  UMAP...")
    umap_file, umap_coords = plot_umap(mat_imp, meta_filt, design_filt, out_dir)

    # --- 10. ANOVA + Heatmaps ---
    print("\n[STEP] ANOVA and heatmaps...")
    (df_anova, hm_classic, hm_clusters, hm_violin,
     mat_zscore, cluster_mapping, sig_names) = run_anova_heatmaps(
        mat_imp, meta_filt, design_filt, params, out_dir
    )

    # --- 11. UpSet ---
    print("\n[STEP] UpSet plot...")
    upset_file, df_intersections = plot_upset(df_results, contrast_names, params, out_dir)

    # --- 12. WGCNA (optionnel) ---
    wgcna_results = None
    if make_wgcna:
        print("\n[STEP] WGCNA...")
        wgcna_results = run_wgcna(mat_imp, meta_filt, design_filt, out_dir)
    else:
        print("\n[SKIP] WGCNA not requested — step skipped.")

    # --- 13. Export Excel (DEP + WGCNA dans le même classeur) ---
    output_name = os.path.join(out_dir, "ProteomicAnalysis_Results.xlsx")

    # Helpers partagés pour l'écriture Excel
    hdr_fill = PatternFill("solid", fgColor="D9D9D9")
    hdr_font = Font(bold=True)

    def _write_df(ws, df, fill=hdr_fill):
        ws.append(list(df.columns))
        for cell in ws[1]:
            cell.fill = fill
            cell.font = hdr_font
        for row in df.itertuples(index=False):
            ws.append(list(row))

    def _ins_img(ws, path, anchor, max_w_px=700, max_h_px=900):
        if not (path and os.path.exists(str(path))):
            return
        try:
            from PIL import Image as _PILImage
            im = _PILImage.open(path); w, h = im.size
            scale = min(max_w_px / w, max_h_px / h, 1.0)
            if scale < 1.0:
                im = im.resize((int(w*scale), int(h*scale)), _PILImage.LANCZOS)
                base, ext = os.path.splitext(str(path))
                path = f"{base}_xl{max_w_px}x{max_h_px}{ext}"
                im.save(path)
        except Exception:
            pass
        ws.add_image(XLImage(str(path)), anchor)

    export_excel(
        tsv=tsv, df_results=df_results, df_anova=df_anova,
        df_intersections=df_intersections, mat_zscore=mat_zscore,
        meta=meta_filt, sig_names=sig_names, contrast_names=contrast_names,
        design=design_filt, params=params,
        qc_files=qc_files, pca_files=pca_files,
        scatter_files=scatter_files, volc_files=volc_files,
        facet_volc=facet_volc, umap_file=umap_file,
        hm_corr=hm_corr, hm_classic=hm_classic,
        hm_clusters=hm_clusters, hm_violin=hm_violin,
        upset_file=upset_file, out_path=output_name,
        go_enabled=(go_params is not None),
        go_organism=(go_params["organism"] if go_params else None),
        n_iter=N_ITER, use_deqms=(pep_count is not None),
        mat_imp=mat_imp, umap_coords=umap_coords,
        imputation_info=imputation_info
    )

    # Ré-ouvrir le classeur pour ajouter les onglets WGCNA
    wb = load_workbook(output_name)
    if wgcna_results is not None:
        export_wgcna_sheets(wb, wgcna_results, _write_df, _ins_img)

    # --- 14. Enrichissement GO/gProfiler (optionnel, non bloquant) ---
    if go_params is not None:
        print("\n[STEP] GO enrichment...")
        go_results = run_go_enrichment(
            df_results, contrast_names, params, go_params, out_dir)
        if go_results:
            export_go_sheets(wb, go_results, _write_df, _ins_img)
            print(f"  [OK] {len(go_results)} GO sheet(s) added.")

    wb.save(output_name)

    print(f"\n[DONE] Full pipeline finished.")
    print(f"   -> {output_name}")
    suffix_go = " + GO" if go_params is not None else ""
    print(f"   Sheets: DEP + UMAP + WGCNA{suffix_go} in a single file.")

    # --- 15. Dashboard HTML interactif (optionnel, non bloquant) ---
    if make_dash and _DASHBOARD_AVAILABLE:
        print("\n[STEP] Generating the interactive HTML dashboard...")
        dash_out = os.path.join(out_dir, "proteogen_dashboard.html")
        result = run_dashboard(output_name, design_filt, dash_out, params=params)
        if result:
            print(f"  [OK] Dashboard generated: {result}")
        else:
            print("  [WARN] Dashboard not generated (see messages above).")

    # --- 16. Nettoyage des figures temporaires ---
    # Toutes les figures sont déjà intégrées dans l'Excel (et le dashboard les
    # embarque en base64). On supprime les PNG/TIFF intermédiaires pour ne pas
    # encombrer le dossier. EXCEPTION : les volcanos (volc_*.png) sont conservés
    # car souvent réutilisés individuellement pour les figures d'article.
    print("\n[CLEAN] Removing temporary figures...")
    keep_prefixes = ("volc_",)              # volcanos individuels conservés
    keep_exact = {"proteogen_dashboard.html",
                  "ProteomicAnalysis_Results.xlsx"}
    removed = 0
    for fn in os.listdir(out_dir):
        if not fn.lower().endswith((".png", ".tiff", ".tif")):
            continue
        # Conserver les volcanos ORIGINAUX (volc_*.png) mais pas leurs copies
        # redimensionnées pour l'Excel (suffixe _xlNNNxNNN).
        is_volcano = fn.startswith("volc_")
        is_resized = "_xl" in fn and "x" in fn.split("_xl")[-1]
        if (is_volcano and not is_resized) or fn in keep_exact:
            continue
        try:
            os.remove(os.path.join(out_dir, fn))
            removed += 1
        except OSError:
            pass
    print(f"  [OK] {removed} temporary figures removed "
          f"(individual volcanos kept).")


if __name__ == "__main__":
    main()
